"""Report generation -- docs/13_Reports.md, driven by the
`report_generation` task (docs/11_BackgroundWorkers.md §8).

`export` is asynchronous by design: a multi-month workbook can't be
built inside a WhatsApp webhook's response window. One `report_jobs`
row carries the status, so the follow-up message and any future API
polling read the same source rather than each tracking progress
separately.
"""

from __future__ import annotations

import dataclasses
import datetime
import decimal
import uuid
from pathlib import Path
from typing import Any

from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from backend.api.formatting import fmt_date
from backend.core.config import get_settings
from backend.core.exceptions import NotFoundError, ValidationError
from backend.core.logging import get_logger
from backend.models import (
    Product,
    PurchaseHeader,
    ReportJob,
    SalesHeader,
    SalesLine,
    Supplier,
    User,
)
from backend.reports.excel.purchase_sheet_template import (
    build_purchase_sheet,
)
from backend.reports.excel.statement_template import StatementEntry, build_statement
from backend.reports.excel.styling import (
    MONEY_FORMAT,
    QTY_FORMAT,
    WARN_FILL,
    autosize,
    write_header,
    write_row,
)
from backend.repositories.accounting_repository import business_today
from backend.repositories.party_repository import OPEN_SALE_STATUSES

logger = get_logger(__name__)
ZERO = decimal.Decimal("0")

REPORT_TYPES = ("purchases", "sales", "stock", "statement", "invoice", "ledger", "cashbook")
LINK_EXPIRY_DAYS = 7


@dataclasses.dataclass(frozen=True)
class GeneratedReport:
    job_id: uuid.UUID
    status: str
    message: str
    notify_number: str | None
    #: Set when the report exists on disk, so the notifier can deliver
    #: the file itself. Telling someone a filename they have no way to
    #: reach is not a report they received.
    file_path: Path | None = None


class ReportService:
    def __init__(self, session: AsyncSession) -> None:
        self._session = session

    async def enqueue(
        self,
        actor: User,
        *,
        report_type: str,
        start: datetime.date,
        end: datetime.date,
        filters: dict[str, Any] | None = None,
    ) -> ReportJob:
        if report_type not in REPORT_TYPES:
            raise ValidationError(
                f"'{report_type}' isn't a report I can export. Try: {', '.join(REPORT_TYPES)}."
            )
        job = ReportJob(
            org_id=actor.org_id,
            report_type=report_type,
            output_format="excel",
            period_start=start,
            period_end=end,
            filters=filters or {},
            status="queued",
            expires_at=datetime.datetime.now(datetime.UTC)
            + datetime.timedelta(days=LINK_EXPIRY_DAYS),
            created_by=actor.id,
        )
        self._session.add(job)
        await self._session.flush()
        return job

    async def generate(self, job_id: uuid.UUID) -> GeneratedReport:
        # Explicit commits rather than `session.begin()` blocks: building
        # a workbook issues reads, which autobegin a transaction, and a
        # later begin() on the same session would raise.
        job = await self._session.get(ReportJob, job_id)
        if job is None:
            raise NotFoundError("report job", str(job_id))

        job.status = "generating"
        await self._session.commit()
        try:
            path, rows = await self._build(job)
        except Exception as exc:  # noqa: BLE001 -- surfaced on the job row
            logger.error("report_generation_failed", job_id=str(job_id), error=str(exc))
            await self._session.rollback()
            job = await self._session.get(ReportJob, job_id)
            assert job is not None
            job.status = "failed"
            job.error = str(exc)[:500]
            await self._session.commit()
            return GeneratedReport(
                job_id=job_id,
                status="failed",
                message=f"❌ That export failed: {exc}",
                notify_number=await self._notify_number(job),
            )

        job.status = "ready"
        job.file_path = str(path)
        job.file_size_bytes = path.stat().st_size
        job.row_count = rows
        await self._session.commit()
        return GeneratedReport(
            job_id=job_id,
            status="ready",
            message=f"📄 Your {job.report_type} export — {rows} row(s).",
            notify_number=await self._notify_number(job),
            file_path=path,
        )

    async def _notify_number(self, job: ReportJob) -> str | None:
        user = await self._session.get(User, job.created_by)
        return user.whatsapp_number if user else None

    def _output_path(self, job: ReportJob) -> Path:
        directory = Path(get_settings().attachments_dir).parent / "reports" / str(job.org_id)
        directory.mkdir(parents=True, exist_ok=True)
        stamp = datetime.datetime.now(datetime.UTC).strftime("%Y%m%d-%H%M%S")
        return directory / f"{job.report_type}-{stamp}-{str(job.id)[:8]}.xlsx"

    async def _build(self, job: ReportJob) -> tuple[Path, int]:
        builders = {
            "purchases": self._build_purchases,
            "sales": self._build_sales,
            "stock": self._build_stock,
            "statement": self._build_statement,
            "ledger": self._build_ledger,
            "cashbook": self._build_cashbook,
            "invoice": self._build_invoice,
        }
        return await builders[job.report_type](job)

    async def _build_purchases(self, job: ReportJob) -> tuple[Path, int]:
        """The legacy-format export (docs/13_Reports.md §5), one sheet per
        bill.

        This selects *which* bills; `DocumentService` builds them. It used
        to build them itself, and the two builders had drifted into
        disagreement: the sheet you downloaded for one bill said MODIFIED
        and carried its corrections, and the sheet for that same bill
        inside "all purchases" said nothing at all. There is now one
        builder, so that cannot recur (docs/28 §3).
        """
        from backend.services.document_service import DocumentService

        header_ids = list(
            (
                await self._session.execute(
                    select(PurchaseHeader.id)
                    .where(
                        PurchaseHeader.org_id == job.org_id,
                        PurchaseHeader.deleted_at.is_(None),
                        PurchaseHeader.status == "confirmed",
                        *self._purchase_scope(job),
                    )
                    .order_by(PurchaseHeader.invoice_date, PurchaseHeader.invoice_no)
                )
            ).scalars()
        )

        bills = [
            built.bill
            for built in await DocumentService(self._session).purchase_bills(job.org_id, header_ids)
        ]
        path = self._output_path(job)
        build_purchase_sheet(bills, title="Purchases").save(path)
        return path, sum(len(bill.rows) for bill in bills)

    def _purchase_scope(self, job: ReportJob) -> list[Any]:
        """Narrow a purchases export to one supplier or one invoice.

        A single invoice ignores the period on purpose: you asked for
        *that bill*, and silently returning nothing because it falls
        outside the default month would be the system being clever at
        your expense.
        """
        invoice_no = job.filters.get("invoice_no")
        if invoice_no:
            return [func.lower(PurchaseHeader.invoice_no) == str(invoice_no).lower()]
        scope: list[Any] = [
            PurchaseHeader.invoice_date >= job.period_start,
            PurchaseHeader.invoice_date <= job.period_end,
        ]
        supplier_id = job.filters.get("supplier_id")
        if supplier_id:
            scope.append(PurchaseHeader.supplier_id == uuid.UUID(str(supplier_id)))
        return scope

    async def _build_invoice(self, job: ReportJob) -> tuple[Path, int]:
        """One bill. Same sheet as the purchases export -- it is the same
        thing, scoped to a single invoice."""
        return await self._build_purchases(job)

    async def _build_ledger(self, job: ReportJob) -> tuple[Path, int]:
        """Every party on one sheet: what they owe, how old it is, and
        when you last did business with them.

        A statement answers "what happened with this one party"; this
        answers "who should I be chasing" -- which is the question you
        open a ledger to ask, and it needs every party side by side.
        """
        from backend.reports.excel.ledger_template import LedgerRow, build_ledger
        from backend.repositories.party_repository import CustomerRepository, SupplierRepository

        role = str(job.filters.get("role", "supplier"))
        if role == "customer":
            parties = await CustomerRepository(self._session).outstanding_parties(job.org_id)
            activity = await self._last_sale_dates(job.org_id)
            heading = "Customers"
        else:
            parties = await SupplierRepository(self._session).outstanding_parties(job.org_id)
            activity = await self._last_purchase_dates(job.org_id)
            heading = "Suppliers"

        today = await business_today(self._session, job.org_id)
        rows = [
            LedgerRow(
                name=party.name,
                outstanding=party.outstanding,
                oldest_date=party.oldest_date,
                days_outstanding=(today - party.oldest_date).days if party.oldest_date else None,
                last_activity=activity.get(party.party_id),
            )
            for party in parties
        ]
        # Each party's own transactions on their own tab. The summary
        # says who to chase; this says what the number is made of.
        histories = {
            party.name: await self._party_entries(job.org_id, role=role, party_id=party.party_id)
            for party in parties
        }
        workbook = build_ledger(rows, heading=heading, as_of=today, histories=histories, role=role)
        path = self._output_path(job)
        workbook.save(path)
        return path, len(rows)

    async def _build_cashbook(self, job: ReportJob) -> tuple[Path, int]:
        """What actually moved through the cash box and the bank, in
        order, with a running balance -- docs/28 §2.4.

        Distinct from `_build_ledger`, which is the *party* ledger. The
        two answer different questions and were only ever called the same
        thing because the partners say "ledger" for both.
        """
        from backend.reports.excel.cashbook_template import CashbookRow, build_cashbook
        from backend.repositories.accounting_repository import LedgerRepository

        repository = LedgerRepository(self._session)
        wanted = str(job.filters.get("account", "")).lower()
        accounts = [wanted] if wanted in {"cash", "bank"} else ["cash", "bank"]
        start = job.period_start or datetime.date.min
        end = job.period_end or datetime.date.max

        by_account: dict[str, list[CashbookRow]] = {}
        counted = 0
        for account in accounts:
            entries = await repository.entries_between(job.org_id, account, start, end)
            cancelled = LedgerRepository.cancelled_ids(entries)
            # Computed over these rows in the order they are about to be
            # printed, rather than read off `resulting_balance`, which
            # runs in insertion order and so disagrees with the page the
            # moment anything is backdated.
            opening = (
                await repository.balance_before(job.org_id, account, start)
                if job.period_start is not None
                else ZERO
            )
            balances = LedgerRepository.running_balances(
                entries, cancelled=cancelled, opening=opening
            )
            by_account[account.capitalize()] = [
                CashbookRow(
                    entry_date=entry.entry_date,
                    entry_type=entry.entry_type.value,
                    details=entry.notes or "",
                    money_in=entry.amount if entry.amount > ZERO else ZERO,
                    money_out=-entry.amount if entry.amount < ZERO else ZERO,
                    balance=balances[entry.id],
                    cancelled=entry.id in cancelled,
                )
                for entry in entries
            ]
            counted += len(entries)

        path = self._output_path(job)
        build_cashbook(by_account, period=(start, end)).save(path)
        return path, counted

    async def _last_purchase_dates(self, org_id: uuid.UUID) -> dict[uuid.UUID, datetime.date]:
        stmt = (
            select(PurchaseHeader.supplier_id, func.max(PurchaseHeader.invoice_date))
            .where(
                PurchaseHeader.org_id == org_id,
                PurchaseHeader.deleted_at.is_(None),
                PurchaseHeader.status == "confirmed",
            )
            .group_by(PurchaseHeader.supplier_id)
        )
        return {row[0]: row[1] for row in (await self._session.execute(stmt)).all()}

    async def _last_sale_dates(self, org_id: uuid.UUID) -> dict[uuid.UUID, datetime.date]:
        from backend.models import SalesHeader

        stmt = (
            select(SalesHeader.customer_id, func.max(SalesHeader.sale_date))
            .where(SalesHeader.org_id == org_id, SalesHeader.deleted_at.is_(None))
            .group_by(SalesHeader.customer_id)
        )
        return {row[0]: row[1] for row in (await self._session.execute(stmt)).all()}

    async def party_entries(
        self,
        org_id: uuid.UUID,
        *,
        role: str,
        party_id: uuid.UUID,
        start: datetime.date | None = None,
        end: datetime.date | None = None,
    ) -> list[StatementEntry]:
        """Public so the dashboard's party ledger reads the same rows
        the exported one does -- two queries for one relationship is two
        places for the balance to disagree."""
        return await self._party_entries(org_id, role=role, party_id=party_id, start=start, end=end)

    async def _party_entries(
        self,
        org_id: uuid.UUID,
        *,
        role: str,
        party_id: uuid.UUID,
        start: datetime.date | None = None,
        end: datetime.date | None = None,
    ) -> list[StatementEntry]:
        """One party's bills/sales and their settlements, merged by time.

        `start`/`end` are optional because a statement is for a period
        while a ledger tab is the whole relationship -- the queries are
        otherwise identical, and two copies of them would be two places
        for the balance to go wrong.
        """
        from backend.models import BankLedger, CashLedger, SalesHeader

        entries: list[StatementEntry] = []
        if role == "supplier":
            stmt = select(PurchaseHeader).where(
                PurchaseHeader.org_id == org_id,
                PurchaseHeader.supplier_id == party_id,
                PurchaseHeader.deleted_at.is_(None),
                PurchaseHeader.status == "confirmed",
            )
            if start is not None:
                stmt = stmt.where(PurchaseHeader.invoice_date >= start)
            if end is not None:
                stmt = stmt.where(PurchaseHeader.invoice_date <= end)
            entries += [
                StatementEntry(
                    at=self._moment(bill.invoice_date, bill.created_at),
                    kind="Purchase",
                    reference=bill.invoice_no,
                    debit=bill.grand_total,
                )
                for bill in (await self._session.execute(stmt)).scalars()
            ]
            source_type, payment_label = "supplier_payment", "Payment"
        else:
            sale_stmt = select(SalesHeader).where(
                SalesHeader.org_id == org_id,
                SalesHeader.customer_id == party_id,
                SalesHeader.deleted_at.is_(None),
                # a cancelled sale was reversed; leaving it in as a debit
                # would make the statement's closing balance disagree
                # with the receivable it is supposed to explain
                SalesHeader.status.in_(OPEN_SALE_STATUSES),
            )
            if start is not None:
                sale_stmt = sale_stmt.where(SalesHeader.sale_date >= start)
            if end is not None:
                sale_stmt = sale_stmt.where(SalesHeader.sale_date <= end)
            entries += [
                StatementEntry(
                    at=self._moment(sale.sale_date, sale.created_at),
                    kind="Sale",
                    reference=str(sale.id)[:8],
                    debit=sale.grand_total,
                )
                for sale in (await self._session.execute(sale_stmt)).scalars()
            ]
            source_type, payment_label = "customer_payment", "Receipt"

        # Reversals land under their own source_type but carry the same
        # party in source_id. Leaving them out was not a cosmetic gap:
        # the settlement they undid was still counted, so six reversed
        # payments made a supplier owed 37,55,350 read as fully settled.
        for ledger in (CashLedger, BankLedger):
            ledger_stmt = select(ledger).where(
                ledger.org_id == org_id,
                ledger.source_type.in_([source_type, "payment_reversal"]),
                ledger.source_id == party_id,
            )
            if start is not None:
                ledger_stmt = ledger_stmt.where(ledger.entry_date >= start)
            if end is not None:
                ledger_stmt = ledger_stmt.where(ledger.entry_date <= end)
            via = "cash" if ledger is CashLedger else "bank"
            # Ledger rows are signed by their effect on *cash*: a payment
            # to a supplier is negative, a receipt from a customer
            # positive. Flipping the supplier side puts both in the same
            # "reduces what is owed" column -- and keeps a reversal, which
            # is the same row with the opposite sign, cancelling its
            # original instead of reading as a second payment.
            direction = 1 if role == "customer" else -1
            entries += [
                StatementEntry(
                    at=self._moment(row.entry_date, row.created_at),
                    kind=(
                        f"Reversal ({via})"
                        if row.source_type == "payment_reversal"
                        else f"{payment_label} ({via})"
                    ),
                    reference=row.notes or "",
                    credit=direction * row.amount,
                )
                for row in (await self._session.execute(ledger_stmt)).scalars()
            ]
        return entries

    async def _build_statement(self, job: ReportJob) -> tuple[Path, int]:
        """Everything that happened with one party, in order, with a
        running balance (docs/13_Reports.md §5).

        Bills and payments come from different tables, so they are read
        separately and merged by time -- which is also the only way the
        balance column can be right.
        """
        from backend.models import Customer

        supplier_id = job.filters.get("supplier_id")
        customer_id = job.filters.get("customer_id")
        role = "supplier" if supplier_id else "customer"
        party_id = uuid.UUID(str(supplier_id or customer_id))

        if role == "supplier":
            supplier = await self._session.get(Supplier, party_id)
            party_name = supplier.name if supplier else "(unknown)"
        else:
            customer = await self._session.get(Customer, party_id)
            party_name = customer.name if customer else "(unknown)"

        # The one implementation, shared with the ledger export's
        # per-party tabs. This method used to carry a second copy of the
        # same three queries -- which is how it kept counting reversed
        # payments long after the copy next door stopped: Noor Traders's
        # statement showed 1,40,85,350 paid against 89,20,350 purchased.
        entries = await self.party_entries(
            job.org_id,
            role=role,
            party_id=party_id,
            start=job.period_start,
            end=job.period_end,
        )

        # What was already owed when the period began. A July statement
        # without it starts from zero and closes on a number that is not
        # the payable -- true of July alone, and read by everyone as
        # wrong.
        opening = ZERO
        if job.period_start is not None:
            earlier = await self.party_entries(
                job.org_id,
                role=role,
                party_id=party_id,
                end=job.period_start - datetime.timedelta(days=1),
            )
            opening = sum((e.debit - e.credit for e in earlier), ZERO)

        workbook = build_statement(
            entries,
            party=party_name,
            role=role,
            period=f"{fmt_date(job.period_start)} to {fmt_date(job.period_end)}"
            if job.period_start and job.period_end
            else "all time",
            opening=opening,
        )
        path = self._output_path(job)
        workbook.save(path)
        return path, len(entries)

    @staticmethod
    def _moment(day: datetime.date | None, recorded: datetime.datetime | None) -> datetime.datetime:
        """The business date decides the order; the recorded timestamp
        supplies the time of day and breaks ties within a date."""
        if day is None:
            return recorded or datetime.datetime.now(datetime.UTC)
        if recorded is None:
            return datetime.datetime.combine(day, datetime.time.min, tzinfo=datetime.UTC)
        return datetime.datetime.combine(day, recorded.timetz())

    async def _build_sales(self, job: ReportJob) -> tuple[Path, int]:
        from openpyxl import Workbook

        from backend.models import Customer

        stmt = (
            select(SalesLine, SalesHeader, Product, Customer)
            .join(SalesHeader, SalesHeader.id == SalesLine.sales_header_id)
            .join(Product, Product.id == SalesLine.product_id)
            .join(Customer, Customer.id == SalesHeader.customer_id)
            .where(
                SalesHeader.org_id == job.org_id,
                SalesHeader.deleted_at.is_(None),
                SalesHeader.status.in_(["confirmed", "partially_returned"]),
                SalesHeader.sale_date >= job.period_start,
                SalesHeader.sale_date <= job.period_end,
            )
            .order_by(SalesHeader.sale_date, SalesHeader.created_at, SalesLine.line_no)
        )
        # NOTE for the same reason the purchase sheet has one: a line
        # that was partly sent back shows a quantity that is no longer
        # the whole story, and a register that doesn't say so is read as
        # if nothing happened (docs/28 §2.1).
        headers = [
            "DATE",
            "CUSTOMER",
            "CODE",
            "DESCRIPTION",
            "QTY",
            "RATE",
            "AMOUNT",
            "COST",
            "NOTE",
        ]
        workbook = Workbook()
        sheet = workbook.active
        assert sheet is not None
        sheet.title = "Sales"
        write_header(sheet, headers)

        formats = {5: QTY_FORMAT, 6: MONEY_FORMAT, 7: MONEY_FORMAT, 8: MONEY_FORMAT}
        total_amount = ZERO
        total_cost = ZERO
        records = (await self._session.execute(stmt)).all()
        for offset, (line, header, product, customer) in enumerate(records, start=2):
            cost = (line.qty * line.avg_cost_at_sale_time).quantize(decimal.Decimal("0.01"))
            total_amount += line.line_total
            total_cost += cost
            write_row(
                sheet,
                offset,
                [
                    header.sale_date.strftime("%d-%m-%Y"),
                    customer.name,
                    product.code,
                    product.description,
                    float(line.qty),
                    float(line.rate),
                    float(line.line_total),
                    float(cost),
                    (
                        f"{line.returned_qty:f} returned"
                        if line.returned_qty and line.returned_qty > ZERO
                        else ""
                    ),
                ],
                formats=formats,
            )
            if line.returned_qty and line.returned_qty > ZERO:
                for column in range(1, len(headers) + 1):
                    sheet.cell(row=offset, column=column).fill = WARN_FILL
        write_row(
            sheet,
            len(records) + 2,
            ["TOTAL", "", "", "", "", "", float(total_amount), float(total_cost), ""],
            formats=formats,
            bold=True,
        )
        autosize(sheet, headers)
        sheet.freeze_panes = "A2"
        path = self._output_path(job)
        workbook.save(path)
        return path, len(records)

    async def _build_stock(self, job: ReportJob) -> tuple[Path, int]:
        from openpyxl import Workbook

        from backend.models import Inventory

        stmt = (
            select(Product, Inventory)
            .join(Inventory, Inventory.product_id == Product.id)
            .where(
                Product.org_id == job.org_id,
                Product.deleted_at.is_(None),
                Product.is_active.is_(True),
            )
            .order_by(Product.code)
        )
        headers = ["CODE", "DESCRIPTION", "ON HAND", "AVG COST", "STOCK VALUE", "REORDER LEVEL"]
        workbook = Workbook()
        sheet = workbook.active
        assert sheet is not None
        sheet.title = "Stock"
        write_header(sheet, headers)

        formats = {3: QTY_FORMAT, 4: MONEY_FORMAT, 5: MONEY_FORMAT, 6: QTY_FORMAT}
        total_value = ZERO
        records = (await self._session.execute(stmt)).all()
        for offset, (product, inventory) in enumerate(records, start=2):
            value = (inventory.qty_on_hand * inventory.weighted_avg_cost).quantize(
                decimal.Decimal("0.01")
            )
            total_value += value
            write_row(
                sheet,
                offset,
                [
                    product.code,
                    product.description,
                    float(inventory.qty_on_hand),
                    float(inventory.weighted_avg_cost),
                    float(value),
                    float(product.reorder_level) if product.reorder_level is not None else "",
                ],
                formats=formats,
            )
        write_row(
            sheet,
            len(records) + 2,
            ["TOTAL", "", "", "", float(total_value), ""],
            formats=formats,
            bold=True,
        )
        autosize(sheet, headers)
        sheet.freeze_panes = "A2"
        path = self._output_path(job)
        workbook.save(path)
        return path, len(records)
