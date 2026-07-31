"""The purchases export -- docs/13_Reports.md §5.

Column layout matches the partners' existing sheet convention:

    S.NO | QTY | DESCRIPTION | CODE | LABEL | KG | T.KG | RATE | AMOUNT

with a bold totals row summing QTY, T.KG and AMOUNT. That layout is the
*textile product type's* export template, not a hard-coded format --
`COLUMNS` below is the same config-over-code shape `ocr_templates`
uses, so a second product type ships its own column list without this
module's structure changing (docs/00_ProjectVision.md §4).

**LABEL is the brand, not the supplier.** One supplier ships many
brands, and a product code is only unique *within* a brand -- which is
what `products_org_code_active_uq` enforces. A supplier name in that
column implied the wrong thing about which codes may collide.

**One worksheet per bill.** A purchase sheet *is* a bill, and its TOTAL
is that invoice's total. Flattening several bills into one sheet gave a
single total that summed unrelated invoices and belonged to none of
them.

The point of matching the legacy layout is that what comes out looks
like the sheet the partners already read. A golden-file test pins it so
an openpyxl upgrade or a refactor can't quietly drift it.
"""

from __future__ import annotations

import dataclasses
import datetime
import decimal
import re
from typing import Any

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from backend.reports.excel.styling import (
    MONEY_FORMAT,
    QTY_FORMAT,
    WARN_FILL,
    WARN_FONT,
    autosize,
    write_header,
    write_row,
)

ZERO = decimal.Decimal("0")

#: (header, source attribute). Order *is* the sheet's column order.
#:
#: NOTE is appended rather than woven in: the nine columns before it are
#: the layout the partners have read for years, and a tenth on the end
#: is additive. It is blank on every row of an uncorrected bill, so such
#: a sheet still reads exactly as it always did (docs/28 §2.1).
COLUMNS: list[tuple[str, str]] = [
    ("S.NO", "serial"),
    ("QTY", "pieces"),
    ("DESCRIPTION", "description"),
    ("CODE", "code"),
    ("LABEL", "label"),
    ("KG", "weight_per_unit"),
    ("T.KG", "total_weight"),
    ("RATE", "rate"),
    ("AMOUNT", "amount"),
    ("NOTE", "note"),
]

#: 1-indexed columns carrying a quantity, shown to 3dp.
NUMERIC_COLUMNS = (2, 6, 7)
#: Money columns, in the same en-IN grouping every WhatsApp reply uses.
MONEY_COLUMNS = (8, 9)
#: Summed in the totals row. KG and RATE are per-unit figures; adding
#: them up would produce a number that means nothing.
TOTALLED_COLUMNS = (2, 7, 9)

#: Excel forbids these characters in a sheet name and truncates at 31.
_SHEET_NAME_BANNED = re.compile(r"[\[\]:*?/\\]")
_SHEET_NAME_LIMIT = 31


@dataclasses.dataclass(frozen=True)
class PurchaseSheetRow:
    serial: int
    pieces: decimal.Decimal | None
    description: str
    code: str
    #: The brand. Blank when the purchase never recorded one -- an empty
    #: cell reads as "not recorded", which is true, and cannot later be
    #: mistaken for a brand actually called that.
    label: str
    weight_per_unit: decimal.Decimal | None
    total_weight: decimal.Decimal | None
    rate: decimal.Decimal | None = None
    amount: decimal.Decimal | None = None
    #: What changed on *this* row since the bill was confirmed, e.g.
    #: "Received 800.000 → 960.000 (30-07)". Blank on untouched rows.
    #: Without it a corrected quantity sits in the table looking exactly
    #: like a billed one, and the reader has no reason to scroll to the
    #: CHANGES block that would have told them otherwise.
    note: str = ""


@dataclasses.dataclass(frozen=True)
class PurchaseBill:
    """One invoice -- what the partners call one purchase sheet."""

    supplier: str
    invoice_no: str
    invoice_date: datetime.date | None
    rows: list[PurchaseSheetRow]
    #: Extra caption lines under the heading -- freight, the payment
    #: that settled it, whatever the document is about.
    notes: list[str] = dataclasses.field(default_factory=list)
    #: Every change made to this bill since it was confirmed, as
    #: "when · who · what". A corrected sheet that does not say it was
    #: corrected is worse than no sheet: two copies in circulation and
    #: nothing on either saying which is current.
    history: list[str] = dataclasses.field(default_factory=list)
    #: The one-line warning printed above the column headers when this
    #: bill has been changed. A bill is read top-down and stops at the
    #: TOTAL, so a notice that lives below it is not a notice.
    banner: str = ""

    def caption(self) -> str:
        date = self.invoice_date.strftime("%d-%m-%Y") if self.invoice_date else "date not recorded"
        return f"Supplier: {self.supplier}    Invoice: {self.invoice_no}    Date: {date}"


def sheet_name(bill: PurchaseBill, taken: set[str]) -> str:
    """Sheet names are short, restricted, and must be unique -- two
    suppliers can easily both issue an "001". A collision gets a suffix
    rather than silently overwriting the earlier tab."""
    base = _SHEET_NAME_BANNED.sub("-", f"{bill.invoice_no} {bill.supplier}").strip()
    base = (base or "Purchases")[:_SHEET_NAME_LIMIT]
    candidate, suffix = base, 2
    while candidate.lower() in taken:
        tail = f" ({suffix})"
        candidate = base[: _SHEET_NAME_LIMIT - len(tail)] + tail
        suffix += 1
    taken.add(candidate.lower())
    return candidate


def build_purchase_sheet(bills: list[PurchaseBill], *, title: str = "Purchases") -> Workbook:
    """One workbook, one worksheet per bill.

    An export covering no purchases still produces a readable empty
    sheet -- a workbook with zero sheets is a file Excel refuses to open.
    """
    workbook = Workbook()
    default = workbook.active
    assert default is not None
    workbook.remove(default)

    if not bills:
        bills = [PurchaseBill(supplier="", invoice_no=title, invoice_date=None, rows=[])]

    taken: set[str] = set()
    for bill in bills:
        _write_bill(workbook.create_sheet(sheet_name(bill, taken)), bill)
    return workbook


def _write_bill(sheet: Worksheet, bill: PurchaseBill) -> None:
    headers = [header for header, _ in COLUMNS]
    # The bill's identity sits above its table: with one tab per invoice,
    # a truncated tab name alone can't tell two of them apart.
    write_row(sheet, 1, [bill.caption(), *[""] * (len(COLUMNS) - 1)], bold=True)

    # A bill that was changed says so before the reader reaches a single
    # number. Nineteen corrections printed below the TOTAL were read, in
    # practice, as no corrections at all (docs/28 §1.1).
    header_row = 2
    if bill.banner:
        write_row(sheet, 2, [bill.banner, *[""] * (len(COLUMNS) - 1)])
        for index in range(1, len(COLUMNS) + 1):
            cell = sheet.cell(row=2, column=index)
            cell.fill = WARN_FILL
            cell.font = WARN_FONT
        header_row = 3

    write_header(sheet, headers, row=header_row)

    formats: dict[int, str] = {index: QTY_FORMAT for index in NUMERIC_COLUMNS}
    formats.update({index: MONEY_FORMAT for index in MONEY_COLUMNS})

    first_data_row = header_row + 1
    for offset, row in enumerate(bill.rows, start=first_data_row):
        write_row(
            sheet,
            offset,
            [_cell_value(row, attribute) for _, attribute in COLUMNS],
            formats=formats,
        )
        # Tinting the whole row, not just its note: at 52 lines the eye
        # finds a band of colour and would never find one filled cell.
        if row.note:
            for index in range(1, len(COLUMNS) + 1):
                sheet.cell(row=offset, column=index).fill = WARN_FILL

    total_row = len(bill.rows) + first_data_row
    totals: list[Any] = [""] * len(COLUMNS)
    totals[0] = "TOTAL"
    for index in TOTALLED_COLUMNS:
        attribute = COLUMNS[index - 1][1]
        totals[index - 1] = float(sum((getattr(row, attribute) or ZERO for row in bill.rows), ZERO))
    write_row(sheet, total_row, totals, formats=formats, bold=True)

    cursor = total_row + 2
    for note in bill.notes:
        write_row(sheet, cursor, [note, *[""] * (len(COLUMNS) - 1)])
        cursor += 1
    if bill.history:
        cursor += 1
        write_row(sheet, cursor, ["CHANGES", *[""] * (len(COLUMNS) - 1)], bold=True)
        for entry in bill.history:
            cursor += 1
            write_row(sheet, cursor, [entry, *[""] * (len(COLUMNS) - 1)])

    autosize(sheet, headers)
    sheet.freeze_panes = f"A{first_data_row}"


def _cell_value(row: PurchaseSheetRow, attribute: str) -> Any:
    value = getattr(row, attribute)
    if isinstance(value, decimal.Decimal):
        # openpyxl writes Decimal as a string; float here is presentation
        # only -- every stored and computed figure stays Decimal.
        return float(value)
    return value
