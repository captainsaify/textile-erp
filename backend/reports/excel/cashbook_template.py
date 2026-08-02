"""The cash & bank book export -- docs/28 §2.4.

    DATE | TYPE | DETAILS | IN | OUT | BALANCE | STATUS

Not to be confused with `ledger_template.py`, which is the *party*
ledger -- who owes what and for how long. This one is the cashbook: what
actually moved through the cash box and the bank account, in order, with
a running balance. It is the sheet the dashboard's Ledger tab shows, and
until now it was the one view of the business with no export at all.

**Reversed rows stay, and are excluded from the totals.** Nothing is
ever deleted (`CLAUDE.md` rule 3), so a reversed payment and the entry
that reversed it are both real rows and both belong on the page. They do
not belong in the money-in and money-out totals, where between them they
claimed twice an amount that never moved -- which is what made the
dashboard read ₹2cr when the business had moved half that.
"""

from __future__ import annotations

import dataclasses
import datetime
import decimal

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from backend.reports.excel.styling import (
    MONEY_FORMAT,
    WARN_FILL,
    autosize,
    write_header,
    write_row,
)

ZERO = decimal.Decimal("0")

HEADERS = ["DATE", "TYPE", "DETAILS", "IN", "OUT", "BALANCE", "STATUS"]
MONEY_COLUMNS = (4, 5, 6)


@dataclasses.dataclass(frozen=True)
class CashbookRow:
    entry_date: datetime.date
    entry_type: str
    details: str
    money_in: decimal.Decimal
    money_out: decimal.Decimal
    balance: decimal.Decimal
    #: True for a reversal and for the entry it reversed. Both are shown;
    #: neither is counted.
    cancelled: bool = False


def build_cashbook(
    rows_by_account: dict[str, list[CashbookRow]],
    *,
    period: tuple[datetime.date, datetime.date] | None = None,
) -> Workbook:
    """One worksheet per account.

    Cash and bank on one sheet would give a running balance belonging to
    neither -- the same mistake flattening several bills into one
    purchase sheet made.
    """
    workbook = Workbook()
    default = workbook.active
    assert default is not None
    workbook.remove(default)

    if not rows_by_account:
        rows_by_account = {"Cash": []}

    for account, rows in rows_by_account.items():
        _write_account(workbook.create_sheet(account[:31]), account, rows, period)
    return workbook


def _write_account(
    sheet: Worksheet,
    account: str,
    rows: list[CashbookRow],
    period: tuple[datetime.date, datetime.date] | None,
) -> None:
    caption = f"{account} book"
    if period is not None:
        caption += f"    {period[0].strftime('%d-%m-%Y')} to {period[1].strftime('%d-%m-%Y')}"
    write_row(sheet, 1, [caption, *[""] * (len(HEADERS) - 1)], bold=True)
    write_header(sheet, HEADERS, row=2)

    formats = {index: MONEY_FORMAT for index in MONEY_COLUMNS}
    for offset, row in enumerate(rows, start=3):
        write_row(
            sheet,
            offset,
            [
                row.entry_date.strftime("%d-%m-%Y"),
                row.entry_type,
                row.details,
                float(row.money_in) or "",
                float(row.money_out) or "",
                float(row.balance),
                "CANCELLED" if row.cancelled else "",
            ],
            formats=formats,
        )
        if row.cancelled:
            for index in range(1, len(HEADERS) + 1):
                sheet.cell(row=offset, column=index).fill = WARN_FILL

    live = [row for row in rows if not row.cancelled]
    total_in = sum((row.money_in for row in live), ZERO)
    total_out = sum((row.money_out for row in live), ZERO)
    # The closing balance is what the columns above it add up to. It used
    # to be the last row's own balance, which is a different number the
    # moment a period has an opening balance or a row is cancelled --
    # and it read as the sheet contradicting itself.
    if not rows:
        opening = ZERO
    elif rows[0].cancelled:
        opening = rows[0].balance  # a cancelled row never moved it
    else:
        opening = rows[0].balance - rows[0].money_in + rows[0].money_out
    closing = opening + total_in - total_out
    write_row(
        sheet,
        len(rows) + 3,
        [
            "TOTAL",
            "",
            f"{len(rows) - len(live)} cancelled row(s) excluded",
            float(total_in),
            float(total_out),
            float(closing),
            "",
        ],
        formats=formats,
        bold=True,
    )

    autosize(sheet, HEADERS)
    sheet.freeze_panes = "A3"
