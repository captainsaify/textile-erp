"""Shared workbook styling -- docs/13_Reports.md §5.

Built cell-by-cell with openpyxl rather than `pandas.to_excel`, whose
default styling doesn't preserve the column widths, borders and bold
totals row the partners are used to seeing.
"""

from __future__ import annotations

from typing import Any

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

HEADER_FILL = PatternFill("solid", fgColor="D9D9D9")
HEADER_FONT = Font(bold=True)
TOTAL_FONT = Font(bold=True)
THIN = Side(style="thin", color="808080")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

#: en-IN digit grouping (₹1,23,456.00) -- the convention every WhatsApp
#: reply already uses, so the spreadsheet doesn't read as a different
#: system's output. docs/13_Reports.md §5.
MONEY_FORMAT = '[>=10000000]"₹"##\\,##\\,##\\,##0.00;[>=100000]"₹"##\\,##\\,##0.00;"₹"##,##0.00'
QTY_FORMAT = "#,##0.000"


def write_header(sheet: Worksheet, headers: list[str], row: int = 1) -> None:
    for index, title in enumerate(headers, start=1):
        cell = sheet.cell(row=row, column=index, value=title)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")


def write_row(
    sheet: Worksheet,
    row: int,
    values: list[Any],
    *,
    formats: dict[int, str] | None = None,
    bold: bool = False,
) -> None:
    for index, value in enumerate(values, start=1):
        cell = sheet.cell(row=row, column=index, value=value)
        cell.border = BORDER
        if bold:
            cell.font = TOTAL_FONT
        if formats and index in formats:
            cell.number_format = formats[index]


def autosize(sheet: Worksheet, headers: list[str], minimum: int = 9) -> None:
    """Width from the widest rendered value, so nothing shows as ####."""
    for index in range(1, len(headers) + 1):
        letter = get_column_letter(index)
        widest = max(
            (len(str(cell.value)) for cell in sheet[letter] if cell.value is not None),
            default=minimum,
        )
        sheet.column_dimensions[letter].width = max(minimum, min(widest + 3, 60))
