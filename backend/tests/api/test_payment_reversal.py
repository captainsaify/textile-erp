"""Reversing a payment -- docs/25_PaymentReversals.md.

A settlement does two things: it moves money, and it marks bills as
settled. Reversing only the first leaves bills showing paid that nobody
paid -- the payable understated, which is the direction that loses money
without anyone noticing. Both are unwound in one transaction.
"""

from __future__ import annotations

import datetime
import decimal
import uuid
from collections.abc import AsyncIterator

import pytest
import sqlalchemy as sa
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker

from backend.core.exceptions import NotFoundError, ValidationError
from backend.models import PurchaseHeader, Supplier, User
from backend.services.settlement_service import PaymentReversalService, SettlementService
from backend.tests.conftest import (
    SEEDED_MAIN_WAREHOUSE_ID,
    SEEDED_ORG_ID,
    purge_business_rows,
)

D = decimal.Decimal
ORG = uuid.UUID(SEEDED_ORG_ID)


@pytest.fixture(autouse=True)
async def clean(session_factory: async_sessionmaker[AsyncSession]) -> AsyncIterator[None]:
    yield
    await purge_business_rows(session_factory)


async def _bill(
    session_factory: async_sessionmaker[AsyncSession], actor: User, *, total: str = "10000"
) -> tuple[str, str]:
    suffix = uuid.uuid4().hex[:5]
    async with session_factory() as session:
        supplier = Supplier(org_id=ORG, name=f"Wagdia {suffix}", created_by=actor.id)
        session.add(supplier)
        await session.flush()
        session.add(
            PurchaseHeader(
                org_id=ORG,
                supplier_id=supplier.id,
                warehouse_id=uuid.UUID(SEEDED_MAIN_WAREHOUSE_ID),
                invoice_no=f"INV-{suffix}",
                invoice_date=datetime.date.today(),
                subtotal=D(total),
                grand_total=D(total),
                status="confirmed",
                created_by=actor.id,
            )
        )
        await session.commit()
    return f"Wagdia {suffix}", f"INV-{suffix}"


async def _pay(
    session_factory: async_sessionmaker[AsyncSession], actor: User, supplier: str, amount: str
) -> str:
    async with session_factory() as session:
        await SettlementService(session).pay_supplier(
            actor, supplier_name=supplier, amount=D(amount), via="cash", against=None
        )
    async with session_factory() as session:
        row = (
            await session.execute(
                sa.text(
                    "SELECT id FROM audit_logs WHERE action = 'payment.paid' "
                    "ORDER BY created_at DESC LIMIT 1"
                )
            )
        ).first()
        assert row is not None
        return str(row[0])[:8]


async def test_reversing_takes_the_money_back_off_the_bill_too(
    session_factory: async_sessionmaker[AsyncSession], staff_user: User
) -> None:
    """The part that matters: the bill reopens. Reversing only the ledger
    would leave it marked settled by money that has been taken back."""
    supplier, invoice = await _bill(session_factory, staff_user, total="10000")
    reference = await _pay(session_factory, staff_user, supplier, "4000")

    async with session_factory() as session:
        paid_before = (
            await session.execute(sa.text("SELECT amount_paid FROM purchase_headers"))
        ).scalar_one()
        assert paid_before == D("4000.00")

    async with session_factory() as session:
        result = await PaymentReversalService(session).reverse(staff_user, reference=reference)

    assert result.kind == "paid"
    assert result.amount == D("4000.00")
    assert invoice in " ".join(result.unapplied)

    async with session_factory() as session:
        paid_after, status = (
            await session.execute(
                sa.text("SELECT amount_paid, payment_status FROM purchase_headers")
            )
        ).one()
        assert paid_after == D("0.00")
        assert status == "unpaid"
        # and the payable is whole again
        assert result.outstanding_after == D("10000.00")


async def test_the_cash_comes_back_and_the_books_balance(
    session_factory: async_sessionmaker[AsyncSession], staff_user: User
) -> None:
    supplier, _ = await _bill(session_factory, staff_user)
    reference = await _pay(session_factory, staff_user, supplier, "4000")

    async with session_factory() as session:
        cash_before = (
            await session.execute(sa.text("SELECT sum(amount) FROM cash_ledger"))
        ).scalar_one()

    async with session_factory() as session:
        await PaymentReversalService(session).reverse(staff_user, reference=reference)

    async with session_factory() as session:
        cash_after = (
            await session.execute(sa.text("SELECT sum(amount) FROM cash_ledger"))
        ).scalar_one()
        assert cash_after == cash_before + D("4000.00")

        debits, credits = (
            await session.execute(sa.text("SELECT sum(debit), sum(credit) FROM journal_lines"))
        ).one()
        assert debits == credits


async def test_a_payment_cannot_be_reversed_twice(
    session_factory: async_sessionmaker[AsyncSession], staff_user: User
) -> None:
    """Otherwise the money comes back as many times as you ask."""
    supplier, _ = await _bill(session_factory, staff_user)
    reference = await _pay(session_factory, staff_user, supplier, "4000")

    async with session_factory() as session:
        await PaymentReversalService(session).reverse(staff_user, reference=reference)
    async with session_factory() as session:
        with pytest.raises(ValidationError, match="already been reversed"):
            await PaymentReversalService(session).reverse(staff_user, reference=reference)


async def test_a_partial_payment_leaves_the_bill_partial(
    session_factory: async_sessionmaker[AsyncSession], staff_user: User
) -> None:
    """Two payments, one reversed: the other must still count."""
    supplier, _ = await _bill(session_factory, staff_user, total="10000")
    first = await _pay(session_factory, staff_user, supplier, "3000")
    await _pay(session_factory, staff_user, supplier, "2000")

    async with session_factory() as session:
        await PaymentReversalService(session).reverse(staff_user, reference=first)

    async with session_factory() as session:
        paid, status = (
            await session.execute(
                sa.text("SELECT amount_paid, payment_status FROM purchase_headers")
            )
        ).one()
        assert paid == D("2000.00"), "the payment that wasn't reversed still stands"
        assert status == "partial"


async def test_an_unknown_payment_is_named(
    session_factory: async_sessionmaker[AsyncSession], staff_user: User
) -> None:
    async with session_factory() as session:
        with pytest.raises(NotFoundError):
            await PaymentReversalService(session).reverse(staff_user, reference="deadbeef")


def test_payment_is_offered_and_routed_like_the_other_reversals() -> None:
    from backend.api.commands import wizards

    assert "payment" in wizards.REVERSIBLE_ENTITIES
    assert wizards.WIZARDS["delete"].reroute({"entity": "payment", "reference": "ab12cd34"}) == (
        "undo",
        "payment ab12cd34",
    )


def test_a_bill_reference_asks_for_a_reference_not_a_product_code() -> None:
    """ "Give its code or name, e.g. TRP" is right for a product and
    wrong for everything else the same slot serves."""
    from backend.api.commands.wizards import _reference_wording

    for entity in ("purchase", "sale", "expense", "payment"):
        question, example = _reference_wording({"entity": entity})
        assert f"Which {entity}?" in question
        assert "TRP" not in example

    question, example = _reference_wording({"entity": "product"})
    assert "code or name" in question
    assert "TRP" in example


# --------------------------------------------------------------------
# undoing a bill that still has money on it
# --------------------------------------------------------------------


async def test_undoing_a_paid_bill_asks_what_to_do_with_the_money(
    session_factory: async_sessionmaker[AsyncSession], staff_user: User
) -> None:
    """Both silent answers are wrong: reversing the payment takes back
    money that really was sent, and leaving it orphans a receipt against
    a bill that no longer exists."""
    from backend.api.command_types import RequestContext
    from backend.api.commands.correction_commands import handle_undo
    from backend.api.interactive import Buttons
    from backend.services.session_service import (
        AWAITING_UNDO_PAYMENT_CHOICE,
        SessionService,
    )

    supplier, invoice = await _bill(session_factory, staff_user, total="10000")
    await _pay(session_factory, staff_user, supplier, "4000")
    ctx = RequestContext(user=staff_user, session_factory=session_factory)

    result = await handle_undo(f"purchase {invoice}", ctx)

    assert isinstance(result.interactive, Buttons)
    assert "4,000" in result.reply
    assert [c.title for c in result.interactive.choices] == [
        "Reverse both",
        "Keep the money",
        "Cancel",
    ]
    state = await SessionService(session_factory).get(ORG, staff_user.id)
    assert state.state == AWAITING_UNDO_PAYMENT_CHOICE
    assert state.context["reference"] == invoice


async def test_cancelling_that_question_changes_nothing(
    session_factory: async_sessionmaker[AsyncSession], staff_user: User
) -> None:
    from backend.api.command_types import RequestContext
    from backend.api.commands import undo_payment_choice
    from backend.api.commands.correction_commands import handle_undo
    from backend.services.session_service import IDLE, SessionService

    supplier, invoice = await _bill(session_factory, staff_user, total="10000")
    await _pay(session_factory, staff_user, supplier, "4000")
    ctx = RequestContext(user=staff_user, session_factory=session_factory)
    await handle_undo(f"purchase {invoice}", ctx)

    state = await SessionService(session_factory).get(ORG, staff_user.id)
    result = await undo_payment_choice.handle_choice("undo cancel", ctx, state)

    assert "Left everything as it was" in result.reply
    assert (await SessionService(session_factory).get(ORG, staff_user.id)).state == IDLE
    async with session_factory() as session:
        paid = (
            await session.execute(sa.text("SELECT amount_paid FROM purchase_headers"))
        ).scalar_one()
        assert paid == D("4000.00"), "the payment is untouched"


async def test_an_unpaid_bill_is_undone_without_the_question(
    session_factory: async_sessionmaker[AsyncSession], staff_user: User
) -> None:
    """The question only exists because money is involved."""
    from backend.api.command_types import RequestContext
    from backend.api.commands.correction_commands import handle_undo

    _, invoice = await _bill(session_factory, staff_user, total="10000")
    ctx = RequestContext(user=staff_user, session_factory=session_factory)

    result = await handle_undo(f"purchase {invoice}", ctx)

    assert result.interactive is None


# --------------------------------------------------------------------
# backdating -- docs/06_Accounting.md, ledgers copied from a paper book
# --------------------------------------------------------------------


async def test_a_payment_can_be_filed_under_the_day_it_actually_moved(
    owner_user: User, session_factory: async_sessionmaker[AsyncSession]
) -> None:
    """A month of payments entered in one sitting all landed on today,
    which put every one of them in the wrong cash-flow period."""
    from backend.api.command_types import RequestContext
    from backend.api.commands.settlement_commands import handle_paid
    from backend.models import CashLedger

    supplier, _ = await _bill(session_factory, owner_user, total="600000")
    ctx = RequestContext(user=owner_user, session_factory=session_factory, message_id="m-date")

    result = await handle_paid(f"{supplier} 600000 cash on 28-07-2026", ctx)

    assert "Payment made" in result.reply
    async with session_factory() as session:
        entry_date = (
            await session.execute(sa.select(CashLedger.entry_date).where(CashLedger.org_id == ORG))
        ).scalar_one()
    assert entry_date == datetime.date(2026, 7, 28)


async def test_a_payment_dated_in_the_future_is_refused(
    owner_user: User, session_factory: async_sessionmaker[AsyncSession]
) -> None:
    from backend.api.command_types import RequestContext
    from backend.api.commands.settlement_commands import handle_paid

    supplier, _ = await _bill(session_factory, owner_user)
    ctx = RequestContext(user=owner_user, session_factory=session_factory, message_id="m-future")
    tomorrow = (datetime.date.today() + datetime.timedelta(days=400)).strftime("%d-%m-%Y")

    result = await handle_paid(f"{supplier} 5000 cash on {tomorrow}", ctx)

    assert "in the future" in result.reply


async def test_a_supplier_whose_name_holds_on_keeps_it(
    owner_user: User, session_factory: async_sessionmaker[AsyncSession]
) -> None:
    """`on` is only a date clause when a date follows it -- otherwise
    "Delivery on Time" would lose two words out of its name."""
    from backend.api.commands.settlement_commands import parse_settlement

    command = parse_settlement("Delivery on Time Traders 5000 cash", "paid")

    assert command.party == "Delivery on Time Traders"
    assert command.on is None


async def test_a_reversed_payment_cancels_itself_out_of_the_party_ledger(
    owner_user: User, session_factory: async_sessionmaker[AsyncSession]
) -> None:
    """A reversal is a compensating entry with the opposite sign. Taking
    the absolute value of every ledger row turned it into a *second*
    payment, so six reversed payments made a supplier who was owed
    37,55,350 read as -51,65,000."""
    from backend.services.report_service import ReportService
    from backend.services.settlement_service import PaymentReversalService

    supplier, _ = await _bill(session_factory, owner_user, total="10000")
    reference = await _pay(session_factory, owner_user, supplier, "4000")
    async with session_factory() as session:
        supplier_id = (
            await session.execute(sa.select(Supplier.id).where(Supplier.name == supplier))
        ).scalar_one()

    async with session_factory() as session:
        entries = await ReportService(session).party_entries(
            ORG, role="supplier", party_id=supplier_id
        )
    assert sum((e.debit - e.credit for e in entries), D("0")) == D("6000")

    async with session_factory() as session:
        await PaymentReversalService(session).reverse(owner_user, reference=reference)

    async with session_factory() as session:
        entries = await ReportService(session).party_entries(
            ORG, role="supplier", party_id=supplier_id
        )
    # both rows are still there -- nothing is deleted -- and they net off
    assert len(entries) == 3
    assert sum((e.debit - e.credit for e in entries), D("0")) == D("10000")


async def test_a_reversal_and_its_original_are_kept_but_not_counted(
    owner_user: User, session_factory: async_sessionmaker[AsyncSession]
) -> None:
    """Both halves stay in the ledger -- nothing is deleted -- but
    neither is money that moved, and counting them made a month of
    ~1cr of payments read as 2.3cr."""
    from backend.repositories.accounting_repository import LedgerRepository
    from backend.services.settlement_service import PaymentReversalService

    supplier, _ = await _bill(session_factory, owner_user, total="10000")
    reference = await _pay(session_factory, owner_user, supplier, "4000")
    async with session_factory() as session:
        await PaymentReversalService(session).reverse(owner_user, reference=reference)

    async with session_factory() as session:
        entries = await LedgerRepository(session).recent_entries(ORG, "cash", limit=50)
    cancelled = LedgerRepository.cancelled_ids(entries)

    # both rows are there
    assert len(entries) == 2
    # and both are excluded, so the pair contributes nothing either way
    assert len(cancelled) == 2
    counted = [entry for entry in entries if entry.id not in cancelled]
    assert counted == []


async def test_an_untouched_payment_is_still_counted(
    owner_user: User, session_factory: async_sessionmaker[AsyncSession]
) -> None:
    """The exclusion has to be narrow, or a real payment disappears."""
    from backend.repositories.accounting_repository import LedgerRepository

    supplier, _ = await _bill(session_factory, owner_user, total="10000")
    await _pay(session_factory, owner_user, supplier, "4000")

    async with session_factory() as session:
        entries = await LedgerRepository(session).recent_entries(ORG, "cash", limit=50)
    assert LedgerRepository.cancelled_ids(entries) == set()


async def test_the_exported_statement_and_the_payable_agree(
    owner_user: User, session_factory: async_sessionmaker[AsyncSession]
) -> None:
    """`export statement` used to carry its own copy of the same three
    queries, which is how it kept counting reversed payments long after
    the copy next door stopped. A statement whose closing balance
    disagrees with the payable it explains is worse than no statement --
    and the ledger export's summary tab, which reads the payable
    directly, made the disagreement visible in one file."""
    import datetime

    from openpyxl import load_workbook

    from backend.repositories.party_repository import SupplierRepository
    from backend.services.report_service import ReportService
    from backend.services.settlement_service import PaymentReversalService

    supplier, _ = await _bill(session_factory, owner_user, total="10000")
    reference = await _pay(session_factory, owner_user, supplier, "4000")
    async with session_factory() as session:
        await PaymentReversalService(session).reverse(owner_user, reference=reference)
        supplier_id = (
            await session.execute(sa.select(Supplier.id).where(Supplier.name == supplier))
        ).scalar_one()

    async with session_factory() as session:
        payable = await SupplierRepository(session).outstanding(ORG, supplier_id)
    async with session_factory() as session, session.begin():
        job = await ReportService(session).enqueue(
            owner_user,
            report_type="statement",
            start=datetime.date.today() - datetime.timedelta(days=30),
            end=datetime.date.today() + datetime.timedelta(days=1),
            filters={"supplier_id": str(supplier_id)},
        )
        job_id = job.id
    async with session_factory() as session:
        built = await ReportService(session).generate(job_id)

    assert payable == D("10000")
    assert built.file_path is not None
    cells = [
        cell.value
        for row in load_workbook(built.file_path).worksheets[0].iter_rows()
        for cell in row
    ]
    # the reversal is on the statement, named, and the closing balance
    # is the payable
    assert any(isinstance(value, str) and "Reversal" in value for value in cells)
    assert D(str(cells[-1])) == payable


async def test_a_period_statement_opens_with_what_was_already_owed(
    owner_user: User, session_factory: async_sessionmaker[AsyncSession]
) -> None:
    """A July statement for a supplier billed in June started from zero
    and closed on a number that was true of July alone. Everyone reads
    that as wrong, because it is not what they are owed."""
    import datetime

    from openpyxl import load_workbook

    from backend.repositories.party_repository import SupplierRepository
    from backend.services.report_service import ReportService

    supplier, _ = await _bill(session_factory, owner_user, total="10000")
    async with session_factory() as session:
        supplier_id = (
            await session.execute(sa.select(Supplier.id).where(Supplier.name == supplier))
        ).scalar_one()
        # backdate the bill to before the statement period
        await session.execute(
            sa.update(PurchaseHeader)
            .where(PurchaseHeader.supplier_id == supplier_id)
            .values(invoice_date=datetime.date.today() - datetime.timedelta(days=60))
        )
        await session.commit()
    await _pay(session_factory, owner_user, supplier, "4000")

    async with session_factory() as session:
        payable = await SupplierRepository(session).outstanding(ORG, supplier_id)
    async with session_factory() as session, session.begin():
        job = await ReportService(session).enqueue(
            owner_user,
            report_type="statement",
            start=datetime.date.today() - datetime.timedelta(days=7),
            end=datetime.date.today() + datetime.timedelta(days=1),
            filters={"supplier_id": str(supplier_id)},
        )
        job_id = job.id
    async with session_factory() as session:
        built = await ReportService(session).generate(job_id)

    assert built.file_path is not None
    cells = [
        cell.value
        for row in load_workbook(built.file_path).worksheets[0].iter_rows()
        for cell in row
    ]
    assert "Opening balance" in cells
    # the bill is outside the window; the closing balance is still what
    # is owed
    assert D(str(cells[-1])) == payable == D("6000")


async def test_a_payment_carries_its_note_into_the_party_statement(
    owner_user: User, session_factory: async_sessionmaker[AsyncSession]
) -> None:
    """The partners' own book explains half its payments -- "in ac
    mahadev", "through hanif pune". Without somewhere to put that, a
    statement can show ₹1,65,000 moved and not that it moved through
    somebody else."""
    from backend.api.command_types import RequestContext
    from backend.api.commands.settlement_commands import handle_paid, parse_settlement
    from backend.services.report_service import ReportService

    command = parse_settlement("Wagdia 4000 cash note: through Hanif Pune", "paid")
    assert command.party == "Wagdia"
    assert command.note == "through Hanif Pune"

    supplier, _ = await _bill(session_factory, owner_user, total="10000")
    ctx = RequestContext(user=owner_user, session_factory=session_factory, message_id="m-note")
    result = await handle_paid(f"{supplier} 4000 cash note: through Hanif Pune", ctx)
    assert "through Hanif Pune" in result.reply

    async with session_factory() as session:
        supplier_id = (
            await session.execute(sa.select(Supplier.id).where(Supplier.name == supplier))
        ).scalar_one()
        entries = await ReportService(session).party_entries(
            ORG, role="supplier", party_id=supplier_id
        )
    payment = next(entry for entry in entries if "Payment" in entry.kind)
    assert "through Hanif Pune" in payment.reference


async def test_a_payment_survives_its_bill_being_renumbered(
    owner_user: User, session_factory: async_sessionmaker[AsyncSession]
) -> None:
    """The bug this test exists for.

    Allocations used to be resolved back to a bill by (party, invoice
    number). Both halves are mutable -- `erp merge` changes the party,
    `erp fix --invoice-no` changes the number -- so a payment made
    before either operation could no longer be found, and the reversal
    moved the money in the ledger while leaving the bill showing
    settled. Silently. The allocation now carries the bill's id, which
    does not move.
    """
    supplier, invoice = await _bill(session_factory, owner_user, total="10000")
    payment = await _pay(session_factory, owner_user, supplier, "4000")

    async with session_factory() as session:
        header = (
            await session.execute(
                sa.select(PurchaseHeader).where(PurchaseHeader.invoice_no == invoice)
            )
        ).scalar_one()
        assert header.amount_paid == D("4000")
        header.invoice_no = "RENUMBERED-001"
        await session.commit()

    async with session_factory() as session:
        await PaymentReversalService(session).reverse(owner_user, reference=payment)

    async with session_factory() as session:
        header = (
            await session.execute(
                sa.select(PurchaseHeader).where(PurchaseHeader.invoice_no == "RENUMBERED-001")
            )
        ).scalar_one()
        assert header.amount_paid == D("0"), (
            "the reversal did not find the bill and left it showing settled"
        )
        assert header.payment_status == "unpaid"


async def test_an_allocation_whose_bill_vanished_refuses_loudly(
    owner_user: User, session_factory: async_sessionmaker[AsyncSession]
) -> None:
    """An allocation that cannot be matched used to be skipped with
    `continue`. The money came back in the ledger and the bill kept
    showing paid, and nothing said so -- the payable understated, which
    is the direction that loses money quietly. It now refuses."""
    supplier, invoice = await _bill(session_factory, owner_user, total="10000")
    payment = await _pay(session_factory, owner_user, supplier, "4000")

    # An entry in the old format: a reference, and no id to fall back on.
    async with session_factory() as session:
        await session.execute(
            sa.text(
                "UPDATE audit_logs SET after_state = jsonb_set("
                "  after_state, '{allocations}',"
                """  '[{"reference": "GONE-999", "applied": "4000"}]'::jsonb)"""
                " WHERE cast(id as text) LIKE :prefix"
            ),
            {"prefix": f"{payment}%"},
        )
        await session.commit()

    async with session_factory() as session:
        with pytest.raises(ValidationError, match="GONE-999"):
            await PaymentReversalService(session).reverse(owner_user, reference=payment)


# --------------------------------------------------------------------
# editing a payment: reverse and re-record, in one go
# --------------------------------------------------------------------


async def test_editing_the_amount_moves_the_bill_with_it(
    session_factory: async_sessionmaker[AsyncSession], staff_user: User
) -> None:
    """The reason this is not a `UPDATE ledger SET amount`: the bill was
    settled by the old figure. An edit that changed only the ledger would
    leave a ₹10,000 bill showing ₹4,000 paid against a payment of
    ₹6,000."""
    from backend.services.settlement_service import PaymentEditService

    supplier, _ = await _bill(session_factory, staff_user, total="10000")
    reference = await _pay(session_factory, staff_user, supplier, "4000")

    async with session_factory() as session:
        result = await PaymentEditService(session).edit(
            staff_user, reference=reference, amount=D("6000")
        )

    assert result.kind == "paid"
    assert result.old_amount == D("4000.00")
    assert result.new_amount == D("6000.00")
    assert result.reference != result.old_reference

    async with session_factory() as session:
        paid, status = (
            await session.execute(
                sa.text("SELECT amount_paid, payment_status FROM purchase_headers")
            )
        ).one()
    assert paid == D("6000.00")
    assert status == "partial"
    assert result.outstanding_after == D("4000.00")


async def test_editing_leaves_the_cash_where_the_new_amount_says(
    session_factory: async_sessionmaker[AsyncSession], staff_user: User
) -> None:
    """Three ledger rows -- original, reversal, correction -- and the
    balance is the one the corrected payment implies, not the sum of a
    muddle."""
    from backend.repositories.accounting_repository import LedgerRepository
    from backend.services.settlement_service import PaymentEditService

    supplier, _ = await _bill(session_factory, staff_user, total="10000")
    async with session_factory() as session:
        before = await LedgerRepository(session).balance(ORG, "cash")

    reference = await _pay(session_factory, staff_user, supplier, "4000")
    async with session_factory() as session:
        await PaymentEditService(session).edit(staff_user, reference=reference, amount=D("2500"))

    async with session_factory() as session:
        after = await LedgerRepository(session).balance(ORG, "cash")
        rows = (await session.execute(sa.text("SELECT count(*) FROM cash_ledger"))).scalar_one()
    assert after == before - D("2500.00")
    assert rows == 3, "the original, the reversal and the correction all stay"


async def test_an_edit_that_changes_nothing_is_refused(
    session_factory: async_sessionmaker[AsyncSession], staff_user: User
) -> None:
    from backend.core.exceptions import ValidationError
    from backend.services.settlement_service import PaymentEditService

    supplier, _ = await _bill(session_factory, staff_user, total="10000")
    reference = await _pay(session_factory, staff_user, supplier, "4000")

    async with session_factory() as session:
        with pytest.raises(ValidationError, match="Nothing about that payment"):
            await PaymentEditService(session).edit(
                staff_user, reference=reference, amount=D("4000")
            )


async def test_an_already_reversed_payment_cannot_be_edited(
    session_factory: async_sessionmaker[AsyncSession], staff_user: User
) -> None:
    """Editing it would re-record money that was deliberately taken back
    -- an undo quietly undone."""
    from backend.core.exceptions import ValidationError
    from backend.services.settlement_service import PaymentEditService

    supplier, _ = await _bill(session_factory, staff_user, total="10000")
    reference = await _pay(session_factory, staff_user, supplier, "4000")
    async with session_factory() as session:
        await PaymentReversalService(session).reverse(staff_user, reference=reference)

    async with session_factory() as session:
        with pytest.raises(ValidationError, match="already reversed"):
            await PaymentEditService(session).edit(
                staff_user, reference=reference, amount=D("5000")
            )
