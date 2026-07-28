"""A second product type must need rows, not code.

`CLAUDE.md`'s last acceptance criterion: adding e.g. "hardware" should
require a `product_types` row, an `ocr_templates` row and a `units`
seed -- no core code changes. HANDOFF.md §7 flagged the export as the
most likely place for a textile assumption to have leaked in, since it
was written last and to the partners' textile layout.

This test adds a hardware type sold by the piece and drives it through
the same purchase, inventory and export paths textile uses. If a
textile assumption is load-bearing anywhere, this fails.
"""

from __future__ import annotations

import datetime
import decimal
import uuid
from collections.abc import AsyncIterator

import pytest
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker

from backend.api.command_types import RequestContext
from backend.models import (
    Inventory,
    Product,
    ProductType,
    PurchaseHeader,
    PurchaseLine,
    Supplier,
    Unit,
    User,
)
from backend.models.enums import UnitKind
from backend.services.inventory_service import InventoryService
from backend.services.report_service import ReportService
from backend.tests.conftest import (
    SEEDED_MAIN_WAREHOUSE_ID,
    SEEDED_ORG_ID,
    purge_business_rows,
)

D = decimal.Decimal
ORG = uuid.UUID(SEEDED_ORG_ID)
WAREHOUSE = uuid.UUID(SEEDED_MAIN_WAREHOUSE_ID)


@pytest.fixture(autouse=True)
async def clean(session_factory: async_sessionmaker[AsyncSession]) -> AsyncIterator[None]:
    """`units` and `product_types` are seed tables and deliberately not
    in `_PURGE_ORDER` -- purging them globally would delete the seeded
    textile type every other test relies on. So this test removes only
    the rows it added, or it leaves a second unit and product type
    behind and unrelated tests start finding two where they expect one
    (HANDOFF.md §5)."""
    import sqlalchemy as sa

    yield
    await purge_business_rows(session_factory)
    async with session_factory() as session:
        # matched by an unmistakable prefix: 'PCS%' would also match the
        # *seeded* PCS unit and delete it out from under every other test
        await session.execute(sa.text("DELETE FROM product_types WHERE code LIKE 'TESTTYPE%'"))
        await session.execute(sa.text("DELETE FROM units WHERE code LIKE 'TESTUNIT%'"))
        await session.commit()


@pytest.fixture
def ctx(staff_user: User, session_factory: async_sessionmaker[AsyncSession]) -> RequestContext:
    return RequestContext(user=staff_user, session_factory=session_factory)


async def test_a_hardware_product_type_needs_only_rows(
    ctx: RequestContext, session_factory: async_sessionmaker[AsyncSession]
) -> None:
    suffix = uuid.uuid4().hex[:6]

    # 1. the rows a new product type is supposed to need
    async with session_factory() as session, session.begin():
        piece = Unit(org_id=ORG, code=f"TESTUNIT{suffix}", name="Pieces", kind=UnitKind.COUNT)
        session.add(piece)
        await session.flush()
        hardware = ProductType(
            org_id=ORG,
            code=f"TESTTYPE{suffix}",
            name="Hardware",
            default_unit_id=piece.id,
        )
        session.add(hardware)
        await session.flush()
        product = Product(
            org_id=ORG,
            product_type_id=hardware.id,
            code=f"BOLT{suffix.upper()}",
            description="M8 Hex Bolt",
            unit_id=piece.id,
            created_by=ctx.user.id,
        )
        session.add(product)
        await session.flush()
        product_id = product.id
        supplier = Supplier(org_id=ORG, name=f"Hardware Co {suffix}", created_by=ctx.user.id)
        session.add(supplier)
        await session.flush()

        header = PurchaseHeader(
            org_id=ORG,
            supplier_id=supplier.id,
            warehouse_id=WAREHOUSE,
            invoice_no=f"HW-{suffix}",
            invoice_date=datetime.date.today(),
            grand_total=D("5000"),
            status="confirmed",
            created_by=ctx.user.id,
        )
        session.add(header)
        await session.flush()
        # sold by the piece: no weight at all, where textile always has one
        session.add(
            PurchaseLine(
                org_id=ORG,
                purchase_header_id=header.id,
                line_no=1,
                product_id=product_id,
                description="M8 Hex Bolt",
                qty=D("500"),
                weight_kg=None,
                total_weight_kg=None,
                rate=D("10"),
                line_total=D("5000"),
                landed_cost_per_unit=D("10"),
            )
        )
        await session.flush()

    # 2. inventory: the same weighted-average path, no textile branch
    async with session_factory() as session, session.begin():
        await InventoryService(session).record_purchase_movement(
            ORG,
            product_id=product_id,
            warehouse_id=WAREHOUSE,
            qty=D("500"),
            landed_cost_per_unit=D("10"),
            source_id=uuid.uuid4(),
            created_by=ctx.user.id,
        )

    async with session_factory() as session:
        from sqlalchemy import select

        row = (
            await session.execute(
                select(Inventory).where(Inventory.org_id == ORG, Inventory.product_id == product_id)
            )
        ).scalar_one()
        assert row.qty_on_hand == D("500")
        assert row.weighted_avg_cost == D("10")

    # 3. the export, which is where a textile assumption would surface
    async with session_factory() as session, session.begin():
        job = await ReportService(session).enqueue(
            ctx.user,
            report_type="purchases",
            start=datetime.date.today(),
            end=datetime.date.today(),
        )
        job_id = job.id
    async with session_factory() as session:
        report = await ReportService(session).generate(job_id)

    assert report.status == "ready", report.message
    assert report.file_path is not None
    sheet = load_workbook(report.file_path).active
    assert sheet is not None

    # the row is present and its money is right; the weight columns are
    # simply blank, which is what "this type has no weight" looks like
    assert sheet.cell(row=3, column=4).value == f"BOLT{suffix.upper()}"
    assert sheet.cell(row=3, column=6).value is None
    assert sheet.cell(row=3, column=8).value == 10
    assert sheet.cell(row=3, column=9).value == 5000
