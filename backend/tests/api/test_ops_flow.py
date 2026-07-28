"""Reconciliation, export and backup -- docs/11_BackgroundWorkers.md
§5/§6/§8, docs/03_Inventory.md §6, docs/13_Reports.md §5.

The reconciliation tests are the ones that matter most: they check
CLAUDE.md's standing acceptance criterion (qty_on_hand equals the signed
sum of movements) against real rows, and they check that a detected
mismatch is *reported and not repaired* — a job that quietly fixed the
number would destroy the only evidence that something upstream is
broken.
"""

from __future__ import annotations

import datetime
import decimal
import uuid
from collections.abc import AsyncIterator
from pathlib import Path
from typing import Any

import pytest
import sqlalchemy as sa
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker

from backend.api.command_types import RequestContext
from backend.api.commands.ops_commands import handle_backup, handle_export, handle_restore
from backend.core.config import get_settings
from backend.models import (
    Inventory,
    Product,
    PurchaseHeader,
    PurchaseLine,
    ReconciliationRun,
    ReportJob,
    Supplier,
    User,
)
from backend.reports.excel.purchase_sheet_template import COLUMNS, build_purchase_sheet
from backend.services.inventory_service import InventoryService
from backend.services.reconciliation_service import ReconciliationService
from backend.services.report_service import ReportService
from backend.tests.conftest import (
    SEEDED_KG_UNIT_ID,
    SEEDED_MAIN_WAREHOUSE_ID,
    SEEDED_ORG_ID,
    SEEDED_TEXTILE_TYPE_ID,
    purge_business_rows,
)

D = decimal.Decimal
ORG = uuid.UUID(SEEDED_ORG_ID)
WAREHOUSE = uuid.UUID(SEEDED_MAIN_WAREHOUSE_ID)


@pytest.fixture(autouse=True)
async def clean(session_factory: async_sessionmaker[AsyncSession]) -> AsyncIterator[None]:
    yield
    await purge_business_rows(session_factory)


@pytest.fixture
def ctx(owner_user: User, session_factory: async_sessionmaker[AsyncSession]) -> RequestContext:
    return RequestContext(user=owner_user, session_factory=session_factory, message_id="m1")


async def _product_with_movement(
    session: AsyncSession, actor: User, *, qty: str = "50", landed: str = "160"
) -> tuple[uuid.UUID, str]:
    suffix = uuid.uuid4().hex[:6]
    product = Product(
        org_id=ORG,
        product_type_id=uuid.UUID(SEEDED_TEXTILE_TYPE_ID),
        code=f"REC{suffix.upper()}",
        description="Reconciliation Test Fabric",
        unit_id=uuid.UUID(SEEDED_KG_UNIT_ID),
        created_by=actor.id,
    )
    session.add(product)
    await session.flush()
    supplier = Supplier(org_id=ORG, name=f"Supp {suffix}", created_by=actor.id)
    session.add(supplier)
    await session.flush()
    header = PurchaseHeader(
        org_id=ORG,
        supplier_id=supplier.id,
        warehouse_id=WAREHOUSE,
        invoice_no=f"INV-{suffix}",
        invoice_date=datetime.date.today(),
        grand_total=(D(qty) * D(landed)),
        status="confirmed",
        created_by=actor.id,
    )
    session.add(header)
    await session.flush()
    line = PurchaseLine(
        org_id=ORG,
        purchase_header_id=header.id,
        line_no=1,
        product_id=product.id,
        description="Reconciliation Test Fabric",
        qty=D(qty),
        weight_kg=D("10"),
        total_weight_kg=D(qty),
        rate=D(landed),
        line_total=(D(qty) * D(landed)),
        landed_cost_per_unit=D(landed),
    )
    session.add(line)
    await session.flush()
    await InventoryService(session).record_purchase_movement(
        ORG,
        product_id=product.id,
        warehouse_id=WAREHOUSE,
        qty=D(qty),
        landed_cost_per_unit=D(landed),
        source_id=line.id,
        created_by=actor.id,
    )
    return product.id, product.code


# --------------------------------------------------------------------
# inventory reconciliation
# --------------------------------------------------------------------


async def test_reconciliation_passes_when_the_cache_matches_the_replay(
    ctx: RequestContext, session_factory: async_sessionmaker[AsyncSession]
) -> None:
    async with session_factory() as session, session.begin():
        await _product_with_movement(session, ctx.user)

    async with session_factory() as session, session.begin():
        outcome = await ReconciliationService(session).run(ORG, "inventory")
    assert outcome.ok
    assert outcome.checked == 1


async def test_a_successful_run_is_recorded_not_left_as_silence(
    ctx: RequestContext, session_factory: async_sessionmaker[AsyncSession]
) -> None:
    """docs/11_BackgroundWorkers.md §6.3 -- otherwise "nothing was wrong"
    and "the job never fired" look identical afterwards."""
    async with session_factory() as session, session.begin():
        await _product_with_movement(session, ctx.user)
    async with session_factory() as session, session.begin():
        await ReconciliationService(session).run(ORG, "inventory")

    async with session_factory() as session:
        run = (
            await session.execute(
                sa.select(ReconciliationRun).where(ReconciliationRun.kind == "inventory")
            )
        ).scalar_one()
    assert run.status == "ok"
    assert run.checked_count == 1
    assert run.finished_at is not None


async def test_reconciliation_detects_a_tampered_cache_and_does_not_fix_it(
    ctx: RequestContext, session_factory: async_sessionmaker[AsyncSession]
) -> None:
    """The core promise: detect, never silently correct
    (docs/03_Inventory.md §6)."""
    async with session_factory() as session, session.begin():
        product_id, code = await _product_with_movement(session, ctx.user)

    # simulate the thing this job exists to catch: the cached balance
    # drifting away from the movements that justify it
    async with session_factory() as session, session.begin():
        await session.execute(
            sa.text("UPDATE inventory SET qty_on_hand = 999 WHERE product_id = :pid"),
            {"pid": product_id},
        )

    async with session_factory() as session, session.begin():
        outcome = await ReconciliationService(session).run(ORG, "inventory")

    assert not outcome.ok
    assert outcome.discrepancies[0].subject == code
    assert outcome.discrepancies[0].cached == "999.000"
    assert outcome.discrepancies[0].replayed == "50.000"

    # and the wrong number is still there -- untouched
    async with session_factory() as session:
        qty = (
            await session.execute(
                sa.select(Inventory.qty_on_hand).where(Inventory.product_id == product_id)
            )
        ).scalar_one()
    assert qty == D("999.000")


async def test_mismatch_alert_names_both_numbers_and_says_it_did_nothing(
    ctx: RequestContext, session_factory: async_sessionmaker[AsyncSession]
) -> None:
    async with session_factory() as session, session.begin():
        product_id, code = await _product_with_movement(session, ctx.user)
    async with session_factory() as session, session.begin():
        await session.execute(
            sa.text("UPDATE inventory SET qty_on_hand = 12 WHERE product_id = :pid"),
            {"pid": product_id},
        )
    async with session_factory() as session, session.begin():
        outcome = await ReconciliationService(session).run(ORG, "inventory")

    text = outcome.alert_text()
    assert code in text
    assert "12.000" in text and "50.000" in text
    assert "Not auto-corrected" in text


async def test_mismatch_detail_is_persisted_for_follow_up(
    ctx: RequestContext, session_factory: async_sessionmaker[AsyncSession]
) -> None:
    async with session_factory() as session, session.begin():
        product_id, _ = await _product_with_movement(session, ctx.user)
    async with session_factory() as session, session.begin():
        await session.execute(
            sa.text("UPDATE inventory SET qty_on_hand = 1 WHERE product_id = :pid"),
            {"pid": product_id},
        )
    async with session_factory() as session, session.begin():
        await ReconciliationService(session).run(ORG, "inventory")

    async with session_factory() as session:
        run = (
            await session.execute(
                sa.select(ReconciliationRun).where(ReconciliationRun.status == "mismatch")
            )
        ).scalar_one()
    assert run.mismatch_count == 1
    assert run.details is not None
    assert run.details[0]["replayed"] == "50.000"


# --------------------------------------------------------------------
# ledger reconciliation
# --------------------------------------------------------------------


async def test_ledger_reconciliation_passes_on_clean_books(
    ctx: RequestContext, session_factory: async_sessionmaker[AsyncSession]
) -> None:
    from backend.api.commands.money_commands import handle_expense

    await handle_expense("transport 500 cash", ctx)
    async with session_factory() as session, session.begin():
        outcome = await ReconciliationService(session).run(ORG, "ledger")
    assert outcome.ok


async def test_ledger_reconciliation_catches_a_broken_running_balance(
    ctx: RequestContext, session_factory: async_sessionmaker[AsyncSession]
) -> None:
    from backend.api.commands.money_commands import handle_expense

    await handle_expense("transport 500 cash", ctx)
    async with session_factory() as session, session.begin():
        await session.execute(sa.text("UPDATE cash_ledger SET resulting_balance = -99999"))

    async with session_factory() as session, session.begin():
        outcome = await ReconciliationService(session).run(ORG, "ledger")
    assert not outcome.ok
    assert "cash balance" in outcome.discrepancies[0].subject


# --------------------------------------------------------------------
# the purchases export
# --------------------------------------------------------------------


def test_purchase_sheet_uses_the_partners_column_order() -> None:
    """docs/13_Reports.md §5 -- the legacy layout is the point of this
    export; a refactor must not quietly reorder it."""
    assert [header for header, _ in COLUMNS] == [
        "S.NO",
        "QTY",
        "DESCRIPTION",
        "CODE",
        "LABEL",
        "KG",
        "T.KG",
        # a purchase sheet is a bill, so it carries what was paid
        "RATE",
        "AMOUNT",
    ]


def _row(serial: int, code: str, pieces: str, total: str, label: str = "Wagdia") -> Any:
    from backend.reports.excel.purchase_sheet_template import PurchaseSheetRow

    return PurchaseSheetRow(
        serial=serial,
        pieces=D(pieces),
        description="Men Zipper Jacket",
        code=code,
        label=label,
        weight_per_unit=D("80"),
        total_weight=D(total),
        rate=D("115"),
        amount=D(total) * D("115"),
    )


def test_purchase_sheet_writes_headers_rows_and_a_totals_row() -> None:
    from backend.reports.excel.purchase_sheet_template import PurchaseBill

    rows = [_row(1, "35A", "10", "800"), _row(2, "VVP-1", "19", "1520")]
    bill = PurchaseBill(
        supplier="Wagdia Textiles",
        invoice_no="INV-001",
        invoice_date=datetime.date(2026, 7, 27),
        rows=rows,
    )
    workbook = build_purchase_sheet([bill])
    sheet = workbook.active
    assert sheet is not None

    # row 1 identifies the bill; the table starts at row 2
    assert "INV-001" in str(sheet.cell(row=1, column=1).value)
    assert "Wagdia Textiles" in str(sheet.cell(row=1, column=1).value)
    assert [cell.value for cell in sheet[2]] == [
        "S.NO",
        "QTY",
        "DESCRIPTION",
        "CODE",
        "LABEL",
        "KG",
        "T.KG",
        "RATE",
        "AMOUNT",
    ]
    assert sheet.cell(row=3, column=4).value == "35A"
    assert sheet.cell(row=3, column=8).value == 115

    # totals: QTY, T.KG and AMOUNT summed; KG and RATE are per-unit
    # figures, so summing them would mean nothing
    total_row = len(rows) + 3
    assert sheet.cell(row=total_row, column=1).value == "TOTAL"
    assert sheet.cell(row=total_row, column=2).value == 29
    assert sheet.cell(row=total_row, column=7).value == 2320
    assert sheet.cell(row=total_row, column=9).value == 2320 * 115
    assert sheet.cell(row=total_row, column=6).value == ""
    assert sheet.cell(row=total_row, column=8).value == ""
    assert sheet.cell(row=total_row, column=1).font.bold


def test_each_bill_gets_its_own_sheet_and_its_own_total() -> None:
    """A purchase sheet is a bill. One flat sheet across three invoices
    produced a single TOTAL that summed unrelated bills."""
    from backend.reports.excel.purchase_sheet_template import PurchaseBill

    workbook = build_purchase_sheet(
        [
            PurchaseBill(
                "Wagdia", "INV-001", datetime.date(2026, 7, 27), [_row(1, "35A", "10", "800")]
            ),
            PurchaseBill(
                "Shree", "INV-002", datetime.date(2026, 7, 28), [_row(1, "22D", "19", "1520")]
            ),
        ]
    )

    assert workbook.sheetnames == ["INV-001 Wagdia", "INV-002 Shree"]
    assert workbook["INV-001 Wagdia"].cell(row=4, column=7).value == 800
    assert workbook["INV-002 Shree"].cell(row=4, column=7).value == 1520


def test_two_suppliers_using_the_same_invoice_number_get_separate_sheets() -> None:
    """Excel silently overwrites a duplicate sheet name; an invoice "001"
    is not rare enough to risk that."""
    from backend.reports.excel.purchase_sheet_template import PurchaseBill

    workbook = build_purchase_sheet(
        [
            PurchaseBill("Wagdia", "001", None, [_row(1, "35A", "10", "800")]),
            PurchaseBill("Wagdia", "001", None, [_row(1, "22D", "19", "1520")]),
        ]
    )

    assert len(workbook.sheetnames) == 2
    assert workbook.sheetnames[1].endswith("(2)")


def test_an_empty_period_still_produces_an_openable_file() -> None:
    workbook = build_purchase_sheet([])
    assert len(workbook.sheetnames) == 1


async def test_export_enqueues_a_job_and_reports_the_reference(
    ctx: RequestContext, session_factory: async_sessionmaker[AsyncSession]
) -> None:
    result = await handle_export("purchases month", ctx)
    assert "Building your purchases export" in result.reply

    async with session_factory() as session:
        job = (await session.execute(sa.select(ReportJob))).scalar_one()
    assert job.report_type == "purchases"
    assert job.status == "queued"
    assert str(job.id)[:8] in result.reply


async def test_export_rejects_an_unknown_report(ctx: RequestContext) -> None:
    result = await handle_export("everything month", ctx)
    assert "isn't a report I can export" in result.reply


async def test_generated_purchases_workbook_contains_the_data(
    ctx: RequestContext, session_factory: async_sessionmaker[AsyncSession]
) -> None:
    async with session_factory() as session, session.begin():
        _, code = await _product_with_movement(session, ctx.user)

    async with session_factory() as session, session.begin():
        job = await ReportService(session).enqueue(
            ctx.user,
            report_type="purchases",
            start=datetime.date.today() - datetime.timedelta(days=1),
            end=datetime.date.today() + datetime.timedelta(days=1),
        )
        job_id = job.id
    async with session_factory() as session:
        report = await ReportService(session).generate(job_id)

    assert report.status == "ready"
    async with session_factory() as session:
        stored = await session.get(ReportJob, job_id)
        assert stored is not None
        assert stored.row_count == 1
        workbook = load_workbook(stored.file_path)
    sheet = workbook.active
    assert sheet is not None
    # row 1 names the bill, row 2 is the header, so data starts at row 3
    assert sheet.cell(row=3, column=4).value == code


async def test_a_failed_report_records_the_error_on_the_job(
    ctx: RequestContext, session_factory: async_sessionmaker[AsyncSession]
) -> None:
    """A failure has to be visible on the row both the follow-up message
    and any API poll read, not only in the logs."""
    async with session_factory() as session, session.begin():
        job = await ReportService(session).enqueue(
            ctx.user,
            report_type="purchases",
            start=datetime.date.today(),
            end=datetime.date.today(),
        )
        job_id = job.id
    async with session_factory() as session, session.begin():
        await session.execute(
            sa.text("UPDATE report_jobs SET report_type = 'nonsense' WHERE id = :id"),
            {"id": job_id},
        )

    async with session_factory() as session:
        report = await ReportService(session).generate(job_id)
    assert report.status == "failed"

    async with session_factory() as session:
        stored = await session.get(ReportJob, job_id)
        assert stored is not None
    assert stored.status == "failed"
    assert stored.error


# --------------------------------------------------------------------
# backup / restore guards
# --------------------------------------------------------------------


@pytest.fixture
def isolated_backup_dir(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> Path:
    """BackupService derives its directory from settings.attachments_dir.
    Without redirecting it these tests read the developer's real backups
    and pass or fail depending on what happens to be on disk."""
    from backend.core.config import Settings
    from backend.services import backup_service

    real = get_settings()
    patched = Settings(**{**real.model_dump(), "attachments_dir": str(tmp_path / "attachments")})
    monkeypatch.setattr(backup_service, "get_settings", lambda: patched)
    return tmp_path / "backups"


async def test_backup_listing_when_none_exist(
    ctx: RequestContext, isolated_backup_dir: Path
) -> None:
    result = await handle_backup("", ctx)
    assert "No backups yet" in result.reply


async def test_restore_without_confirmation_explains_the_risk(ctx: RequestContext) -> None:
    """A restore that one mistyped word could trigger is a data-loss
    incident waiting to happen."""
    result = await handle_restore("backup-20260101T000000Z-abc.dump", ctx)
    assert "replaces ALL current data" in result.reply
    assert "confirm" in result.reply


async def test_restore_refuses_a_mismatched_confirmation(ctx: RequestContext) -> None:
    result = await handle_restore("backup-a.dump confirm backup-b.dump", ctx)
    assert "To confirm" in result.reply or "No backup named" in result.reply


async def test_ops_command_permissions() -> None:
    from backend.api.whatsapp_commands import COMMAND_REGISTRY
    from backend.models.enums import UserRole

    assert COMMAND_REGISTRY["export"].min_role == UserRole.STAFF
    assert COMMAND_REGISTRY["backup"].min_role == UserRole.OWNER
    assert COMMAND_REGISTRY["restore"].min_role == UserRole.OWNER


def test_beat_schedule_covers_every_scheduled_task() -> None:
    """A task that exists but is never scheduled is dead code; one that
    is scheduled but doesn't exist fails only at 2am."""
    from backend.workers import tasks as task_module
    from backend.workers.schedule import CELERYBEAT_SCHEDULE

    for entry in CELERYBEAT_SCHEDULE.values():
        name = entry["task"]
        assert hasattr(task_module, name), f"{name} is scheduled but not defined"


def test_the_schedule_is_actually_attached_to_the_celery_app() -> None:
    """`celery beat` reads beat_schedule off the configured app. A
    schedule that is defined but never assigned gives a Beat process
    that starts cleanly and fires nothing — silently, forever."""
    from backend.workers.app import celery_app
    from backend.workers.schedule import CELERYBEAT_SCHEDULE

    assert celery_app.conf.beat_schedule == CELERYBEAT_SCHEDULE
    assert celery_app.conf.beat_schedule, "beat has no entries"


def test_every_scheduled_task_is_registered_with_the_app() -> None:
    """Beat dispatches by task *name*; a name Beat knows but the app
    doesn't produces an unroutable message at 2am, not an import error
    at deploy time."""
    from backend.workers import tasks  # noqa: F401 -- registers the tasks
    from backend.workers.app import celery_app
    from backend.workers.schedule import CELERYBEAT_SCHEDULE

    for entry in CELERYBEAT_SCHEDULE.values():
        assert entry["task"] in celery_app.tasks, f"{entry['task']} not registered"


# --------------------------------------------------------------------
# delivering the report, not just announcing it
# --------------------------------------------------------------------


class RecordingClient:
    """Stands in for the Cloud API client; records what was sent."""

    def __init__(self, *, upload_succeeds: bool = True) -> None:
        self.upload_succeeds = upload_succeeds
        self.texts: list[tuple[str, str]] = []
        self.documents: list[tuple[str, Path, str, str]] = []

    async def send_text(self, to_number: str, body: str) -> bool:
        self.texts.append((to_number, body))
        return True

    async def send_document(
        self, to_number: str, path: Path, *, filename: str, caption: str
    ) -> bool:
        self.documents.append((to_number, path, filename, caption))
        return self.upload_succeeds


async def _finished_report(tmp_path: Path, client: object) -> None:
    from backend.services import whatsapp_client as client_module
    from backend.services.report_service import GeneratedReport
    from backend.workers.tasks import _deliver_report

    workbook = tmp_path / "purchases-20260728-abcdef.xlsx"
    workbook.write_bytes(b"xlsx")
    record = GeneratedReport(
        job_id=uuid.uuid4(),
        status="ready",
        message="📄 Your purchases export — 26 row(s).",
        notify_number="+919000000000",
        file_path=workbook,
    )
    original = client_module.get_whatsapp_client
    client_module.get_whatsapp_client = lambda: client  # type: ignore[assignment,return-value]
    try:
        await _deliver_report(record)
    finally:
        client_module.get_whatsapp_client = original


async def test_a_ready_report_arrives_as_a_file(tmp_path: Path) -> None:
    """The old message named a file inside a container and said it would
    expire -- neither of which the recipient could act on. Nothing was
    ever uploaded."""
    client = RecordingClient()
    await _finished_report(tmp_path, client)

    assert len(client.documents) == 1
    _, path, filename, caption = client.documents[0]
    assert filename == "purchases-20260728-abcdef.xlsx"
    assert path.read_bytes() == b"xlsx"
    assert "26 row(s)" in caption
    assert client.texts == [], "the caption carries the message; a separate text repeats it"


async def test_a_failed_upload_says_so_instead_of_claiming_delivery(tmp_path: Path) -> None:
    client = RecordingClient(upload_succeeds=False)
    await _finished_report(tmp_path, client)

    assert len(client.documents) == 1
    assert len(client.texts) == 1
    assert "couldn't attach the file" in client.texts[0][1]


async def test_a_transport_without_files_falls_back_to_text(tmp_path: Path) -> None:
    """The whatsapp-web.js bridge has no send_document; the flow still
    completes rather than raising (docs/19 §3)."""

    class TextOnlyClient:
        def __init__(self) -> None:
            self.texts: list[tuple[str, str]] = []

        async def send_text(self, to_number: str, body: str) -> bool:
            self.texts.append((to_number, body))
            return True

    client = TextOnlyClient()
    await _finished_report(tmp_path, client)

    assert len(client.texts) == 1
    assert "26 row(s)" in client.texts[0][1]


async def test_label_is_the_brand_so_overlapping_codes_stay_distinguishable(
    ctx: RequestContext, session_factory: async_sessionmaker[AsyncSession]
) -> None:
    """LABEL used to carry the supplier, which implied codes were unique
    per supplier. They are not: a code is unique within a *brand*
    (`products_org_code_active_uq`), and one supplier ships many brands.
    The export has to show which brand a code belongs to, or two
    legitimately different products read as the same row.
    """
    from backend.models import Brand

    suffix = uuid.uuid4().hex[:6]
    shared_code = f"VVP{suffix.upper()}"

    async with session_factory() as session, session.begin():
        supplier = Supplier(org_id=ORG, name=f"One Supplier {suffix}", created_by=ctx.user.id)
        brands = [Brand(org_id=ORG, name=f"{name}{suffix}") for name in ("Alpha", "Beta")]
        session.add_all([supplier, *brands])
        await session.flush()

        header = PurchaseHeader(
            org_id=ORG,
            supplier_id=supplier.id,
            warehouse_id=WAREHOUSE,
            invoice_no=f"BR-{suffix}",
            invoice_date=datetime.date.today(),
            grand_total=D("0"),
            status="confirmed",
            created_by=ctx.user.id,
        )
        session.add(header)
        await session.flush()

        # the same code under two brands: legal, and the reason LABEL matters
        for line_no, brand in enumerate(brands, start=1):
            product = Product(
                org_id=ORG,
                product_type_id=uuid.UUID(SEEDED_TEXTILE_TYPE_ID),
                code=shared_code,
                brand_id=brand.id,
                description=f"{brand.name} Velvet Pant",
                unit_id=uuid.UUID(SEEDED_KG_UNIT_ID),
                created_by=ctx.user.id,
            )
            session.add(product)
            await session.flush()
            session.add(
                PurchaseLine(
                    org_id=ORG,
                    purchase_header_id=header.id,
                    line_no=line_no,
                    product_id=product.id,
                    description=product.description,
                    qty=D("100"),
                    weight_kg=D("10"),
                    total_weight_kg=D("100"),
                    rate=D("115"),
                    line_total=D("11500"),
                    landed_cost_per_unit=D("115"),
                )
            )

    async with session_factory() as session, session.begin():
        job = await ReportService(session).enqueue(
            ctx.user,
            report_type="purchases",
            start=datetime.date.today(),
            end=datetime.date.today(),
        )
        job_id = job.id
    async with session_factory() as session:
        report = await ReportService(session).generate(job_id)

    assert report.status == "ready"
    assert report.file_path is not None
    sheet = load_workbook(report.file_path).active
    assert sheet is not None

    codes = [sheet.cell(row=row, column=4).value for row in (3, 4)]
    labels = [sheet.cell(row=row, column=5).value for row in (3, 4)]
    assert codes == [shared_code, shared_code]
    assert sorted(labels) == [f"Alpha{suffix}", f"Beta{suffix}"], "the brand tells the rows apart"
    assert f"One Supplier {suffix}" not in labels

    # and the bill carries what was paid
    assert sheet.cell(row=3, column=8).value == 115
    assert sheet.cell(row=3, column=9).value == 11500
    assert sheet.cell(row=5, column=9).value == 23000


# --------------------------------------------------------------------
# party statement and single-invoice export
# --------------------------------------------------------------------


def test_statement_carries_a_running_balance() -> None:
    """A list of bills beside a list of payments doesn't answer "what do
    I owe them"; one chronological column with a carried balance does."""
    from backend.reports.excel.statement_template import StatementEntry, build_statement

    def at(day: int, hour: int) -> datetime.datetime:
        return datetime.datetime(2026, 7, day, hour, 30, tzinfo=datetime.UTC)

    entries = [
        StatementEntry(at=at(27, 10), kind="Purchase", reference="INV-001", debit=D("40920")),
        StatementEntry(at=at(28, 9), kind="Payment (cash)", reference="", credit=D("10000")),
        StatementEntry(at=at(28, 15), kind="Purchase", reference="INV-002", debit=D("5000")),
    ]
    sheet = build_statement(entries, party="Wagdia", role="supplier", period="July").active
    assert sheet is not None

    assert [cell.value for cell in sheet[3]] == [
        "DATE",
        "TIME",
        "TYPE",
        "REFERENCE",
        "PURCHASED",
        "PAID",
        "BALANCE",
    ]
    # date and time, because "when" was the question
    assert sheet.cell(row=4, column=1).value == "27-07-2026"
    assert sheet.cell(row=4, column=2).value == "10:30"
    # the balance is carried, not recomputed per row
    assert [sheet.cell(row=row, column=7).value for row in (4, 5, 6)] == [40920, 30920, 35920]

    total_row = len(entries) + 4
    assert sheet.cell(row=total_row, column=5).value == 45920
    assert sheet.cell(row=total_row, column=6).value == 10000
    assert sheet.cell(row=total_row, column=7).value == 35920
    assert "Owed to them" in str(sheet.cell(row=total_row + 1, column=1).value)


def test_statement_orders_by_time_not_by_the_order_rows_were_read() -> None:
    """Bills and payments come from different tables; if the merge kept
    read order the balance column would be arithmetic about nothing."""
    from backend.reports.excel.statement_template import StatementEntry, build_statement

    late = datetime.datetime(2026, 7, 28, 9, 0, tzinfo=datetime.UTC)
    early = datetime.datetime(2026, 7, 27, 9, 0, tzinfo=datetime.UTC)
    sheet = build_statement(
        [
            StatementEntry(at=late, kind="Payment (cash)", reference="", credit=D("100")),
            StatementEntry(at=early, kind="Purchase", reference="INV-1", debit=D("500")),
        ],
        party="Wagdia",
        role="supplier",
        period="July",
    ).active
    assert sheet is not None

    assert sheet.cell(row=4, column=3).value == "Purchase"
    assert [sheet.cell(row=row, column=7).value for row in (4, 5)] == [500, 400]


async def test_exporting_one_invoice_ignores_the_period(
    ctx: RequestContext, session_factory: async_sessionmaker[AsyncSession]
) -> None:
    """You asked for that bill. Returning nothing because it falls
    outside the default month would be the system being clever at your
    expense."""
    async with session_factory() as session, session.begin():
        _, code = await _product_with_movement(session, ctx.user)
        invoice_no = (
            await session.execute(sa.text("SELECT invoice_no FROM purchase_headers LIMIT 1"))
        ).scalar_one()

    async with session_factory() as session, session.begin():
        job = await ReportService(session).enqueue(
            ctx.user,
            report_type="invoice",
            # a period that deliberately excludes the bill
            start=datetime.date(2000, 1, 1),
            end=datetime.date(2000, 1, 2),
            filters={"invoice_no": invoice_no},
        )
        job_id = job.id
    async with session_factory() as session:
        report = await ReportService(session).generate(job_id)

    assert report.status == "ready"
    assert report.file_path is not None
    sheet = load_workbook(report.file_path).active
    assert sheet is not None
    assert sheet.cell(row=3, column=4).value == code
    assert invoice_no in str(sheet.cell(row=1, column=1).value)


async def test_a_supplier_scoped_export_excludes_other_suppliers(
    ctx: RequestContext, session_factory: async_sessionmaker[AsyncSession]
) -> None:
    async with session_factory() as session, session.begin():
        _, mine = await _product_with_movement(session, ctx.user)
        _, theirs = await _product_with_movement(session, ctx.user)
        supplier_id = (
            await session.execute(
                sa.text(
                    "SELECT h.supplier_id FROM purchase_headers h "
                    "JOIN purchase_lines l ON l.purchase_header_id = h.id "
                    "JOIN products p ON p.id = l.product_id WHERE p.code = :code"
                ),
                {"code": mine},
            )
        ).scalar_one()

    async with session_factory() as session, session.begin():
        job = await ReportService(session).enqueue(
            ctx.user,
            report_type="purchases",
            start=datetime.date.today(),
            end=datetime.date.today(),
            filters={"supplier_id": str(supplier_id)},
        )
        job_id = job.id
    async with session_factory() as session:
        report = await ReportService(session).generate(job_id)

    assert report.file_path is not None
    workbook = load_workbook(report.file_path)
    codes = {
        sheet.cell(row=row, column=4).value
        for sheet in workbook.worksheets
        for row in range(3, sheet.max_row + 1)
    }
    assert mine in codes
    assert theirs not in codes


def test_the_ledger_ages_the_debt_not_just_its_size() -> None:
    """₹50,000 owed for ninety days is a different problem from ₹50,000
    owed since Tuesday. The ledger is opened to decide who to chase, so
    it sorts by size and marks by age."""
    from backend.reports.excel.ledger_template import LedgerRow, build_ledger

    today = datetime.date(2026, 7, 29)
    rows = [
        LedgerRow(
            name="Small but ancient",
            outstanding=D("5000"),
            oldest_date=datetime.date(2026, 1, 1),
            days_outstanding=209,
            last_activity=datetime.date(2026, 1, 1),
        ),
        LedgerRow(
            name="Large and fresh",
            outstanding=D("500000"),
            oldest_date=datetime.date(2026, 7, 28),
            days_outstanding=1,
            last_activity=datetime.date(2026, 7, 28),
        ),
    ]
    sheet = build_ledger(rows, heading="Suppliers", as_of=today).active
    assert sheet is not None

    assert [cell.value for cell in sheet[2]] == [
        "PARTY",
        "OUTSTANDING",
        "OLDEST",
        "DAYS",
        "LAST ACTIVITY",
        "STATUS",
    ]
    # largest first
    assert sheet.cell(row=3, column=1).value == "Large and fresh"
    assert sheet.cell(row=3, column=6).value == "current"
    assert sheet.cell(row=4, column=6).value == "overdue"
    # and the total is the sum, not the top row
    assert sheet.cell(row=5, column=2).value == 505000


def test_a_party_never_traded_with_reads_never_rather_than_blank() -> None:
    from backend.reports.excel.ledger_template import LedgerRow, build_ledger

    sheet = build_ledger(
        [
            LedgerRow(
                name="Opening balance only",
                outstanding=D("1000"),
                oldest_date=None,
                days_outstanding=None,
                last_activity=None,
            )
        ],
        heading="Suppliers",
        as_of=datetime.date(2026, 7, 29),
    ).active
    assert sheet is not None
    assert sheet.cell(row=3, column=5).value == "never"
    assert sheet.cell(row=3, column=6).value == ""
