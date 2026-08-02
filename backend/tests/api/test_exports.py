"""Downloadable sheets from every page -- docs/28 §2.4.

The property under test: a browser download and a WhatsApp `export` run
the *same* builder and write the *same* `report_jobs` row. Two paths to
the same figures is how a dashboard and a chat come to disagree, and an
export that leaves no trace is a copy of the business's numbers walking
out of the building unrecorded.
"""

from __future__ import annotations

import datetime
import decimal
import io
import uuid
from collections.abc import AsyncIterator

import pytest
import sqlalchemy as sa
from httpx import ASGITransport, AsyncClient
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker

from backend.core.security import hash_password
from backend.main import create_app
from backend.models import CashLedger, ReportJob, Supplier, User
from backend.models.enums import LedgerEntryType, UserRole
from backend.tests.conftest import SEEDED_ORG_ID, purge_business_rows

D = decimal.Decimal
ORG = uuid.UUID(SEEDED_ORG_ID)
PASSWORD = "correct-horse-battery-staple"

XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


@pytest.fixture(autouse=True)
async def clean(session_factory: async_sessionmaker[AsyncSession]) -> AsyncIterator[None]:
    yield
    await purge_business_rows(session_factory)


@pytest.fixture
async def client(
    session_factory: async_sessionmaker[AsyncSession], monkeypatch: pytest.MonkeyPatch
) -> AsyncIterator[AsyncClient]:
    import backend.api.deps as deps

    monkeypatch.setattr(deps, "get_session_factory", lambda: session_factory)
    transport = ASGITransport(app=create_app())
    async with AsyncClient(transport=transport, base_url="http://test") as http:
        yield http


@pytest.fixture
async def auth(
    client: AsyncClient, session_factory: async_sessionmaker[AsyncSession]
) -> dict[str, str]:
    async with session_factory() as session, session.begin():
        user = User(
            org_id=ORG,
            full_name="Export Probe",
            email=f"export-{uuid.uuid4().hex[:6]}@example.com",
            password_hash=hash_password(PASSWORD),
            role=UserRole.OWNER,
        )
        session.add(user)
        await session.flush()
        email = user.email
    response = await client.post("/api/v1/auth/login", json={"email": email, "password": PASSWORD})
    assert response.status_code == 200, response.text
    return {"Authorization": f"Bearer {response.json()['access_token']}"}


def _sheets(payload: bytes) -> list[str]:
    return list(load_workbook(io.BytesIO(payload)).sheetnames)


@pytest.mark.parametrize(
    "path",
    [
        "/api/v1/exports/purchases.xlsx",
        "/api/v1/exports/sales.xlsx",
        "/api/v1/exports/stock.xlsx",
        "/api/v1/exports/parties.xlsx?role=supplier",
        "/api/v1/exports/parties.xlsx?role=customer",
        "/api/v1/exports/cashbook.xlsx",
    ],
)
async def test_every_export_streams_a_workbook(
    client: AsyncClient, auth: dict[str, str], path: str
) -> None:
    """An empty period still produces a readable file: a workbook with
    zero sheets is one Excel refuses to open."""
    response = await client.get(path, headers=auth)
    assert response.status_code == 200, response.text
    assert response.headers["content-type"] == XLSX
    assert response.headers["cache-control"] == "no-store"
    assert "attachment;" in response.headers["content-disposition"]
    assert _sheets(response.content)


async def test_an_export_needs_a_token(client: AsyncClient) -> None:
    response = await client.get("/api/v1/exports/purchases.xlsx")
    assert response.status_code == 401


async def test_a_browser_export_is_as_traceable_as_a_chat_one(
    client: AsyncClient, auth: dict[str, str], session_factory: async_sessionmaker[AsyncSession]
) -> None:
    """An export is a copy of the business's figures leaving the system.
    One taken from the browser must leave the same record as one taken
    from WhatsApp (docs/28 §2.4)."""
    assert (await client.get("/api/v1/exports/stock.xlsx", headers=auth)).status_code == 200

    async with session_factory() as session:
        jobs = (
            (await session.execute(sa.select(ReportJob).where(ReportJob.org_id == ORG)))
            .scalars()
            .all()
        )
    assert [job.report_type for job in jobs] == ["stock"]
    assert jobs[0].status == "ready"
    assert jobs[0].created_by is not None


async def test_a_statement_export_needs_the_party_it_is_about(
    client: AsyncClient, auth: dict[str, str]
) -> None:
    # the app maps request validation to 400 (docs/10_API.md §7)
    assert (await client.get("/api/v1/exports/statement.xlsx", headers=auth)).status_code == 400


async def test_a_statement_export_names_the_party(
    client: AsyncClient, auth: dict[str, str], session_factory: async_sessionmaker[AsyncSession]
) -> None:
    async with session_factory() as session, session.begin():
        actor = (await session.execute(sa.select(User).where(User.org_id == ORG))).scalars().first()
        assert actor is not None
        supplier = Supplier(org_id=ORG, name="Noor Probe", created_by=actor.id)
        session.add(supplier)
        await session.flush()
        supplier_id = supplier.id

    response = await client.get(
        f"/api/v1/exports/statement.xlsx?kind=supplier&party_id={supplier_id}", headers=auth
    )
    assert response.status_code == 200, response.text
    sheet = load_workbook(io.BytesIO(response.content)).worksheets[0]
    assert "Noor Probe" in " ".join(
        str(cell.value) for row in sheet.iter_rows(max_row=4) for cell in row if cell.value
    )


async def test_a_missing_document_is_a_404_not_a_500(
    client: AsyncClient, auth: dict[str, str]
) -> None:
    """The dashboard fetches these on every row click; a 500 on a stale
    id would look like the sheet builder is broken."""
    for path in (
        f"/api/v1/purchases/{uuid.uuid4()}/document",
        f"/api/v1/sales/{uuid.uuid4()}/document",
        "/api/v1/payments/deadbeef/document",
    ):
        assert (await client.get(path, headers=auth)).status_code == 404, path


async def test_the_cashbook_shows_reversals_and_does_not_count_them(
    client: AsyncClient, auth: dict[str, str], session_factory: async_sessionmaker[AsyncSession]
) -> None:
    """Nothing is ever deleted, so a reversed payment and the entry that
    reversed it are both real rows and both belong on the page. Counting
    them made a month of 1cr read as 2cr."""
    source = uuid.uuid4()
    async with session_factory() as session, session.begin():
        actor = (await session.execute(sa.select(User).where(User.org_id == ORG))).scalars().first()
        assert actor is not None
        session.add_all(
            [
                CashLedger(
                    org_id=ORG,
                    entry_date=datetime.date(2026, 7, 9),
                    entry_type=LedgerEntryType.PURCHASE_PAYMENT,
                    amount=D("-100000"),
                    resulting_balance=D("-100000"),
                    notes="paid to probe",
                    source_type="payment",
                    source_id=source,
                    created_by=actor.id,
                ),
                CashLedger(
                    org_id=ORG,
                    entry_date=datetime.date(2026, 7, 10),
                    entry_type=LedgerEntryType.PURCHASE_PAYMENT,
                    amount=D("100000"),
                    resulting_balance=D("0"),
                    notes="reversed: paid to probe",
                    source_type="payment_reversal",
                    source_id=source,
                    created_by=actor.id,
                ),
            ]
        )

    response = await client.get("/api/v1/exports/cashbook.xlsx?account=cash", headers=auth)
    assert response.status_code == 200, response.text
    sheet = load_workbook(io.BytesIO(response.content))["Cash"]

    rows = [[cell.value for cell in row] for row in sheet.iter_rows(min_row=3)]
    body, totals = rows[:-1], rows[-1]
    # both halves are on the page ...
    assert len(body) == 2
    assert all(row[6] == "CANCELLED" for row in body)
    # ... and neither is in the totals
    assert totals[3] == 0 and totals[4] == 0
    assert "2 cancelled row(s) excluded" in str(totals[2])


async def test_a_backdated_row_does_not_break_the_balance_column(
    client: AsyncClient, auth: dict[str, str], session_factory: async_sessionmaker[AsyncSession]
) -> None:
    """A payment to Iqbal Bhai dated 20-07 was entered on 01-08. The
    sheet prints rows in date order but read `resulting_balance`, which
    is written in *insertion* order -- so that row sat between 22-07 and
    18-07 carrying the balance from the end of the book, and the BALANCE
    column stopped reconciling with the IN and OUT columns beside it.

    The balance must be the running total of the rows as printed.
    """
    async with session_factory() as session, session.begin():
        actor = (await session.execute(sa.select(User).where(User.org_id == ORG))).scalars().first()
        assert actor is not None
        session.add_all(
            [
                CashLedger(  # entered first, dated later
                    org_id=ORG,
                    entry_date=datetime.date(2026, 7, 20),
                    entry_type=LedgerEntryType.SALE_RECEIPT,
                    amount=D("1000"),
                    resulting_balance=D("1000"),
                    notes="received early",
                    source_type="customer_payment",
                    source_id=uuid.uuid4(),
                    created_by=actor.id,
                ),
                CashLedger(  # entered second, dated *earlier* -- the trap
                    org_id=ORG,
                    entry_date=datetime.date(2026, 7, 10),
                    entry_type=LedgerEntryType.PURCHASE_PAYMENT,
                    amount=D("-400"),
                    resulting_balance=D("600"),
                    notes="backdated payment",
                    source_type="supplier_payment",
                    source_id=uuid.uuid4(),
                    created_by=actor.id,
                ),
            ]
        )

    response = await client.get("/api/v1/exports/cashbook.xlsx?account=cash", headers=auth)
    assert response.status_code == 200, response.text
    sheet = load_workbook(io.BytesIO(response.content))["Cash"]
    rows = [[cell.value for cell in row] for row in sheet.iter_rows(min_row=3)]
    body, totals = rows[:-1], rows[-1]

    # printed in date order: the backdated payment comes first
    assert [row[0] for row in body] == ["10-07-2026", "20-07-2026"]
    # and the balance column follows *that* order, not the insertion one
    assert body[0][5] == -400.0, "the backdated row kept its insertion-order balance"
    assert body[1][5] == 600.0

    # every row's balance is the previous one plus in, minus out
    running = 0.0
    for row in body:
        running += (row[3] or 0) - (row[4] or 0)
        assert row[5] == running, f"balance column breaks at {row[0]}"

    # and the total row agrees with the columns above it
    assert totals[5] == totals[3] - totals[4] == 600.0
