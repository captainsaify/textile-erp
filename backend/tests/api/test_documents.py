"""A document for every transaction -- docs/27_Documents.md.

The property under test throughout: the sheet is built from the
database when it is asked for, so a bill that was corrected produces
the corrected sheet, and the correction is printed on it with who made
it and when. A stored file would have handed back the superseded
version while the correction lived only in a chat message.
"""

from __future__ import annotations

import datetime
import decimal
import uuid
from collections.abc import AsyncIterator

import pytest
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker

from backend.api.command_types import RequestContext
from backend.models import (
    Inventory,
    Product,
    ProductType,
    PurchaseHeader,
    PurchaseLine,
    Supplier,
    User,
)
from backend.services.document_service import DocumentService
from backend.tests.conftest import (
    SEEDED_MAIN_WAREHOUSE_ID,
    SEEDED_ORG_ID,
    purge_business_rows,
)

D = decimal.Decimal
ORG = uuid.UUID(SEEDED_ORG_ID)


@pytest.fixture(autouse=True)
async def clean(session_factory: async_sessionmaker[AsyncSession]) -> AsyncIterator[None]:
    yield
    await purge_business_rows(session_factory)


@pytest.fixture
def ctx(owner_user: User, session_factory: async_sessionmaker[AsyncSession]) -> RequestContext:
    return RequestContext(user=owner_user, session_factory=session_factory, message_id="doc")


@pytest.fixture
async def stocked(
    owner_user: User, session_factory: async_sessionmaker[AsyncSession]
) -> AsyncIterator[dict[str, uuid.UUID]]:
    import sqlalchemy as sa

    ids: dict[str, uuid.UUID] = {}
    async with session_factory() as session:
        product_type = (
            await session.execute(sa.select(ProductType).where(ProductType.org_id == ORG))
        ).scalar_one()
        product = Product(
            org_id=ORG,
            product_type_id=product_type.id,
            code=f"TRP{uuid.uuid4().hex[:4].upper()}",
            description="Trouser Poly",
            unit_id=product_type.default_unit_id,
            created_by=owner_user.id,
        )
        session.add(product)
        await session.flush()
        ids["TRP"] = product.id
        session.add(
            Inventory(
                org_id=ORG,
                product_id=product.id,
                warehouse_id=uuid.UUID(SEEDED_MAIN_WAREHOUSE_ID),
                qty_on_hand=D("150"),
                weighted_avg_cost=D("100"),
            )
        )
        await session.commit()
    yield ids


def _cells(path: str) -> list[str]:
    sheet = load_workbook(path).worksheets[0]
    return [str(cell.value) for row in sheet.iter_rows() for cell in row if cell.value is not None]


async def _bill(
    session_factory: async_sessionmaker[AsyncSession], actor: User, product_id: uuid.UUID
) -> uuid.UUID:
    suffix = uuid.uuid4().hex[:5]
    async with session_factory() as session:
        supplier = Supplier(org_id=ORG, name=f"Docs {suffix}", created_by=actor.id)
        session.add(supplier)
        await session.flush()
        header = PurchaseHeader(
            org_id=ORG,
            supplier_id=supplier.id,
            warehouse_id=uuid.UUID(SEEDED_MAIN_WAREHOUSE_ID),
            invoice_no=f"DOC-{suffix}",
            invoice_date=datetime.date(2026, 7, 19),
            subtotal=D("8000"),
            grand_total=D("8000"),
            status="confirmed",
            created_by=actor.id,
        )
        session.add(header)
        await session.flush()
        session.add(
            PurchaseLine(
                org_id=ORG,
                purchase_header_id=header.id,
                line_no=1,
                product_id=product_id,
                description="Jogging Pant",
                qty=D("80"),
                weight_kg=D("8"),
                rate=D("100"),
                line_total=D("8000"),
            )
        )
        await session.commit()
        return header.id


async def test_the_document_carries_the_bill_as_it_stands_now(
    ctx: RequestContext,
    stocked: dict[str, uuid.UUID],
    session_factory: async_sessionmaker[AsyncSession],
) -> None:
    header_id = await _bill(session_factory, ctx.user, stocked["TRP"])

    async with session_factory() as session:
        document = await DocumentService(session).purchase(ORG, header_id)
    cells = _cells(str(document.path))

    assert "Jogging Pant" in cells
    assert "Grand total: 8,000.00" in cells
    assert "Outstanding: 8,000.00" in cells
    # QTY and KG are derived: a line stores total weight and per-unit
    # weight, and the sheet's QTY column is the bale count
    assert "10" in cells and "8" in cells


async def test_a_correction_is_printed_on_the_document_with_who_and_when(
    ctx: RequestContext,
    stocked: dict[str, uuid.UUID],
    session_factory: async_sessionmaker[AsyncSession],
) -> None:
    """A corrected sheet that does not say it was corrected is worse
    than no sheet: two copies in circulation, neither saying which is
    current."""
    from backend.services.audit_service import AuditService

    header_id = await _bill(session_factory, ctx.user, stocked["TRP"])
    async with session_factory() as session, session.begin():
        await AuditService(session).record(
            ORG,
            ctx.user.id,
            action="purchase.rate_corrected",
            entity_type="purchase_headers",
            entity_id=header_id,
            before_state={"rate": "100"},
            after_state={"rate": "107"},
        )

    async with session_factory() as session:
        document = await DocumentService(session).purchase(ORG, header_id)
    cells = _cells(str(document.path))

    assert "CHANGES" in cells
    change = next(cell for cell in cells if "Rate corrected" in cell)
    assert ctx.user.full_name in change
    assert "100 → 107" in change


async def _line_id(
    session_factory: async_sessionmaker[AsyncSession], header_id: uuid.UUID
) -> uuid.UUID:
    import sqlalchemy as sa

    async with session_factory() as session:
        return (
            await session.execute(
                sa.select(PurchaseLine.id).where(PurchaseLine.purchase_header_id == header_id)
            )
        ).scalar_one()


async def _correct_receipt(
    session_factory: async_sessionmaker[AsyncSession],
    actor: User,
    line_id: uuid.UUID,
    *,
    code: str,
    before: str,
    after: str,
) -> None:
    from backend.services.audit_service import AuditService

    async with session_factory() as session, session.begin():
        await AuditService(session).record(
            ORG,
            actor.id,
            action="purchase.receipt_corrected",
            entity_type="purchase_lines",
            entity_id=line_id,
            before_state={"code": code, "pieces": before, "qty": str(D(before) * 80)},
            after_state={"code": code, "pieces": after, "qty": str(D(after) * 80)},
        )


def _note_column(path: str) -> list[str]:
    """The NOTE cells of the data rows, in order."""
    sheet = load_workbook(path).worksheets[0]
    header_row = 3 if str(sheet.cell(row=2, column=1).value or "").startswith("⚠") else 2
    notes: list[str] = []
    for index in range(header_row + 1, sheet.max_row + 1):
        if str(sheet.cell(row=index, column=1).value) == "TOTAL":
            break
        notes.append(str(sheet.cell(row=index, column=10).value or ""))
    return notes


async def test_a_corrected_row_says_so_on_the_row_itself(
    ctx: RequestContext,
    stocked: dict[str, uuid.UUID],
    session_factory: async_sessionmaker[AsyncSession],
) -> None:
    """The failure this exists to prevent: nineteen corrections recorded
    against bill 003, every one of them on the sheet, and the partner
    reading it concluded nothing had been recorded -- because the table
    showed the corrected quantity looking exactly like a billed one and
    the evidence sat nine rows below the TOTAL (docs/28 §1.1)."""
    header_id = await _bill(session_factory, ctx.user, stocked["TRP"])
    await _correct_receipt(
        session_factory,
        ctx.user,
        await _line_id(session_factory, header_id),
        code="TRP",
        before="10",
        after="12",
    )

    async with session_factory() as session:
        document = await DocumentService(session).purchase(ORG, header_id)

    assert _note_column(str(document.path))[0].startswith("Received 10 → 12 bales")

    # and the warning is above the table, not below the total
    sheet = load_workbook(str(document.path)).worksheets[0]
    banner = str(sheet.cell(row=2, column=1).value)
    assert banner.startswith("⚠ MODIFIED")
    assert "1 change(s)" in banner
    assert ctx.user.full_name in banner


async def test_a_twice_corrected_row_keeps_both_corrections(
    ctx: RequestContext,
    stocked: dict[str, uuid.UUID],
    session_factory: async_sessionmaker[AsyncSession],
) -> None:
    """ANG on bill 003 went 1360 → 800 → 1040. The last correction is
    not the whole story when someone corrected a correction."""
    header_id = await _bill(session_factory, ctx.user, stocked["TRP"])
    line_id = await _line_id(session_factory, header_id)
    await _correct_receipt(session_factory, ctx.user, line_id, code="TRP", before="17", after="10")
    await _correct_receipt(session_factory, ctx.user, line_id, code="TRP", before="10", after="13")

    async with session_factory() as session:
        document = await DocumentService(session).purchase(ORG, header_id)

    note = _note_column(str(document.path))[0]
    assert note.index("17 → 10") < note.index("10 → 13")


async def test_an_untouched_bill_carries_no_banner_and_no_notes(
    ctx: RequestContext,
    stocked: dict[str, uuid.UUID],
    session_factory: async_sessionmaker[AsyncSession],
) -> None:
    """A warning on every document trains everyone to ignore the row."""
    header_id = await _bill(session_factory, ctx.user, stocked["TRP"])

    async with session_factory() as session:
        document = await DocumentService(session).purchase(ORG, header_id)

    sheet = load_workbook(str(document.path)).worksheets[0]
    assert sheet.cell(row=2, column=1).value == "S.NO"  # headers, not a banner
    assert _note_column(str(document.path)) == [""]


async def test_a_rate_correction_marks_every_code_it_named(
    ctx: RequestContext,
    stocked: dict[str, uuid.UUID],
    session_factory: async_sessionmaker[AsyncSession],
) -> None:
    """A rate correction is recorded once against the header but is
    visible on every line it repriced -- and those are exactly the lines
    whose AMOUNT stopped matching the photographed sheet."""
    import sqlalchemy as sa

    from backend.services.audit_service import AuditService

    header_id = await _bill(session_factory, ctx.user, stocked["TRP"])
    async with session_factory() as session:
        code = (
            await session.execute(sa.select(Product.code).where(Product.id == stocked["TRP"]))
        ).scalar_one()
    async with session_factory() as session, session.begin():
        await AuditService(session).record(
            ORG,
            ctx.user.id,
            action="purchase.rate_corrected",
            entity_type="purchase_headers",
            entity_id=header_id,
            before_state={"rate": "150.0000", "codes": f"{code}, SOMETHINGELSE"},
            after_state={"rate": "107", "codes": f"{code}, SOMETHINGELSE"},
        )

    async with session_factory() as session:
        document = await DocumentService(session).purchase(ORG, header_id)

    assert _note_column(str(document.path))[0].startswith("Rate 150 → 107")


async def test_the_web_view_and_the_workbook_are_one_build(
    ctx: RequestContext,
    stocked: dict[str, uuid.UUID],
    session_factory: async_sessionmaker[AsyncSession],
) -> None:
    """The dashboard draws the sheet and the button beside it downloads
    the sheet. Deriving those separately is how two versions of one bill
    end up in circulation (docs/28 §2.2)."""
    header_id = await _bill(session_factory, ctx.user, stocked["TRP"])
    await _correct_receipt(
        session_factory,
        ctx.user,
        await _line_id(session_factory, header_id),
        code="TRP",
        before="10",
        after="12",
    )

    async with session_factory() as session:
        service = DocumentService(session)
        view = await service.purchase_view(ORG, header_id)
        document = await service.purchase(ORG, header_id)

    sheet = load_workbook(str(document.path)).worksheets[0]
    assert view["columns"] == [str(cell.value) for cell in sheet[3]]
    assert view["banner"] == str(sheet.cell(row=2, column=1).value)
    assert view["rows"][0]["changed"] is True
    assert view["rows"][0]["cells"][9] == _note_column(str(document.path))[0]
    assert view["rows"][0]["cells"][3] == str(sheet.cell(row=4, column=4).value)
    assert view["history"]


async def test_a_reversed_bill_says_so(
    ctx: RequestContext,
    stocked: dict[str, uuid.UUID],
    session_factory: async_sessionmaker[AsyncSession],
) -> None:
    """The document is the current truth about the bill, and "this was
    cancelled" is the most important thing it can say."""
    import sqlalchemy as sa

    header_id = await _bill(session_factory, ctx.user, stocked["TRP"])
    async with session_factory() as session, session.begin():
        await session.execute(
            sa.update(PurchaseHeader)
            .where(PurchaseHeader.id == header_id)
            .values(status="cancelled")
        )

    async with session_factory() as session:
        document = await DocumentService(session).purchase(ORG, header_id)
    assert "STATUS: CANCELLED" in _cells(str(document.path))
