"""The draft as a spreadsheet, and multi-item sales.

docs/24_DraftPreview.md. Confirming 26 lines you scrolled past in a chat
window is how a wrong figure gets saved; the same draft as the .xlsx the
partners already read is checkable. Nothing is posted by looking.
"""

from __future__ import annotations

import datetime
import decimal
import uuid
from collections.abc import AsyncIterator
from pathlib import Path

import pytest
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker

from backend.api.command_types import CommandResult, RequestContext
from backend.api.commands import wizards
from backend.api.commands.draft_preview import handle_sheet
from backend.api.interactive import Buttons
from backend.models import User
from backend.services.purchase_service import Draft, DraftLine
from backend.services.session_service import (
    AWAITING_PURCHASE_CONFIRMATION,
    SessionService,
)
from backend.tests.conftest import SEEDED_ORG_ID, purge_business_rows

D = decimal.Decimal
ORG = uuid.UUID(SEEDED_ORG_ID)


@pytest.fixture(autouse=True)
async def clean(session_factory: async_sessionmaker[AsyncSession]) -> AsyncIterator[None]:
    yield
    await purge_business_rows(session_factory)


@pytest.fixture
def ctx(staff_user: User, session_factory: async_sessionmaker[AsyncSession]) -> RequestContext:
    return RequestContext(user=staff_user, session_factory=session_factory, message_id="m1")


def _draft() -> Draft:
    return Draft(
        supplier_id=None,
        supplier_name="Wagdia Textiles",
        invoice_no="INV-77",
        invoice_date=datetime.date(2026, 7, 29),
        brand_id=None,
        brand_name="TOP",
        lines=[
            DraftLine(
                code="35A",
                qty=D("800"),
                rate=D("150"),
                product_id=None,
                resolved_code=None,
                unit_code="KG",
                description="Men Zipper Jacket",
                pieces=D("10"),
                weight_per_unit=D("80"),
            )
        ],
        freight=D("0"),
        other_charges=D("0"),
        declared_total=None,
    )


async def test_a_waiting_purchase_becomes_a_readable_sheet(ctx: RequestContext) -> None:
    await SessionService(ctx.session_factory).set(
        ORG, ctx.user.id, AWAITING_PURCHASE_CONFIRMATION, _draft().to_context()
    )

    result = await handle_sheet("", ctx)

    assert result.attachment is not None
    assert "Nothing is saved yet" in result.reply
    sheet = load_workbook(result.attachment).active
    assert sheet is not None
    assert "INV-77" in str(sheet.cell(row=1, column=1).value)
    assert sheet.cell(row=3, column=4).value == "35A"
    assert sheet.cell(row=3, column=5).value == "TOP"
    assert sheet.cell(row=3, column=8).value == 150
    assert sheet.cell(row=3, column=9).value == 120000


async def test_looking_at_the_draft_does_not_touch_it(ctx: RequestContext) -> None:
    """The whole point is to look *before* deciding. If `sheet` cleared
    the session, CONFIRM afterwards would have nothing to confirm."""
    sessions = SessionService(ctx.session_factory)
    await sessions.set(ORG, ctx.user.id, AWAITING_PURCHASE_CONFIRMATION, _draft().to_context())

    await handle_sheet("", ctx)

    after = await sessions.get(ORG, ctx.user.id)
    assert after.state == AWAITING_PURCHASE_CONFIRMATION
    assert Draft.from_context(after.context).invoice_no == "INV-77"


async def test_asking_for_a_sheet_with_no_draft_says_so(ctx: RequestContext) -> None:
    result = await handle_sheet("", ctx)

    assert result.attachment is None
    assert "no draft waiting" in result.reply


async def test_the_generated_file_is_a_real_workbook(ctx: RequestContext) -> None:
    await SessionService(ctx.session_factory).set(
        ORG, ctx.user.id, AWAITING_PURCHASE_CONFIRMATION, _draft().to_context()
    )
    result = await handle_sheet("", ctx)

    assert result.attachment is not None
    path = Path(result.attachment)
    assert path.suffix == ".xlsx"
    assert path.stat().st_size > 0


# --------------------------------------------------------------------
# multi-item sales
# --------------------------------------------------------------------


def test_a_sale_collects_items_until_told_to_stop() -> None:
    """`filled` is the only state a wizard has, so the loop lives in it
    rather than a counter somewhere else that could disagree."""
    filled = {"customer": "Ravi", "code": "TRP", "qty": "100", "rate": "150", "more": "add"}

    wizards._bank_item(filled)

    # the item is banked and the per-item slots cleared, so they're asked
    # again -- that is the loop
    assert filled["items"] == "TRP 100 150"
    assert "code" not in filled
    assert "more" not in filled

    filled.update({"code": "MJP", "qty": "40", "rate": "200", "more": "done"})
    wizards._bank_item(filled)

    assert filled["items"] == "TRP 100 150\nMJP 40 200"
    assert filled["more"] == "done", "finishing stays filled, so it isn't asked again"


def test_the_assembled_sale_is_the_grammar_someone_could_have_typed() -> None:
    assembled = wizards.WIZARDS["sale"].assemble(
        {"customer": "Ravi Traders", "items": "TRP 100 150\nMJP 40 200"}
    )

    assert assembled == "Customer: Ravi Traders\nTRP 100 150\nMJP 40 200"


async def test_the_sale_wizard_asks_for_another_item(ctx: RequestContext) -> None:
    from backend.services.session_service import AWAITING_COMMAND_SLOT

    started = await wizards.start(wizards.WIZARDS["sale"], "", ctx)
    assert started is not None

    async def answer(text: str) -> CommandResult:
        state = await SessionService(ctx.session_factory).get(ORG, ctx.user.id)
        return await wizards.handle_reply(text, ctx, state)

    await answer("Ravi Traders")
    await answer("TRP")
    await answer("100")
    result = await answer("150")

    # after the first item it asks, rather than assuming one line
    assert isinstance(result.interactive, Buttons)
    assert "Anything else" in result.interactive.body

    await answer("slot add")
    state = await SessionService(ctx.session_factory).get(ORG, ctx.user.id)
    assert state.state == AWAITING_COMMAND_SLOT
    assert state.context["queue"][0] == "code", "it asks for the next code"
    # the validators quantise, and the sale grammar accepts decimals --
    # what matters is that the item was banked, not its formatting
    assert state.context["filled"]["items"] == "TRP 100.00 150.00"


def test_an_ambiguous_answer_to_anything_else_is_not_taken_as_done() -> None:
    with pytest.raises(Exception, match="Add another"):
        wizards._more_items("maybe")
